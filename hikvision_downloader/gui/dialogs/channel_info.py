# 通道信息表格对话框
import csv
from typing import List, Dict, Optional
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QMessageBox, QHeaderView, QAbstractItemView,
    QAbstractScrollArea, QFileDialog,
)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QColor


class ChannelInfoDialog(QDialog):
    """通道信息表格展示对话框，支持导出Excel/CSV"""

    osd_update_signal = pyqtSignal(list, dict)  # osd_updates, device_config

    def __init__(self, device_name: str, table_data: List[Dict], device_config: dict = None, parent=None):
        super().__init__(parent)
        self.device_name = device_name
        self.table_data = table_data
        self.device_config = device_config or {}
        self.setWindowTitle(f"通道信息 - {device_name}")
        self.setMinimumSize(1400, 600)
        self.resize(1600, 750)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        title_label = QLabel(f"📹 {self.device_name} - 通道信息 ({len(self.table_data)}个通道)")
        title_font = QFont()
        title_font.setPointSize(12)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        online_count = sum(1 for row in self.table_data if row['online'] == '在线')
        stats_label = QLabel(f"在线: {online_count} | 离线: {len(self.table_data) - online_count}")
        stats_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(stats_label)

        # 表格
        self.table = QTableWidget()
        self.table.setColumnCount(16)
        headers = [
            "通道号", "通道名称", "在线状态", "IP地址", "协议", "OSD名称",
            "主码流分辨率", "主码流编码", "主码流码率控制", "主码流码率", "主码流帧率",
            "子码流分辨率", "子码流编码", "子码流码率控制", "子码流码率", "子码流帧率"
        ]
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.table_data))
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(False)

        min_widths = [50, 100, 60, 100, 70, 100, 90, 140, 90, 70, 60, 90, 140, 90, 70, 60]
        for i, width in enumerate(min_widths):
            self.table.setColumnWidth(i, width)

        for row_idx, row_data in enumerate(self.table_data):
            item = QTableWidgetItem(str(row_data['channel_no']))
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 0, item)

            self.table.setItem(row_idx, 1, QTableWidgetItem(row_data['channel_name']))

            online_item = QTableWidgetItem(row_data['online'])
            online_item.setTextAlignment(Qt.AlignCenter)
            if row_data['online'] == '在线':
                online_item.setForeground(QColor(0, 153, 76))
            else:
                online_item.setForeground(QColor(232, 17, 35))
            self.table.setItem(row_idx, 2, online_item)

            self.table.setItem(row_idx, 3, QTableWidgetItem(row_data['ip']))
            self.table.setItem(row_idx, 4, QTableWidgetItem(row_data['protocol']))
            self.table.setItem(row_idx, 5, QTableWidgetItem(row_data['osd_name']))

            for col, key in [(6, 'main_resolution'), (7, 'main_codec')]:
                self.table.setItem(row_idx, col, QTableWidgetItem(row_data[key]))

            for col, key in [(8, 'main_bitrate_mode'), (9, 'main_bitrate'), (10, 'main_fps'),
                             (11, 'sub_resolution'), (12, 'sub_codec'),
                             (13, 'sub_bitrate_mode'), (14, 'sub_bitrate'), (15, 'sub_fps')]:
                item = QTableWidgetItem(row_data[key])
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, col, item)

        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.table.resizeColumnsToContents()

        layout.addWidget(self.table)

        # 按钮区域
        btn_layout = QHBoxLayout()

        export_csv_btn = QPushButton("📄 导出CSV")
        export_csv_btn.setToolTip("导出为CSV格式表格文件")
        export_csv_btn.clicked.connect(self._export_csv)
        btn_layout.addWidget(export_csv_btn)

        export_excel_btn = QPushButton("📊 导出Excel")
        export_excel_btn.setToolTip("导出为Excel格式表格文件（需要安装openpyxl）")
        export_excel_btn.clicked.connect(self._export_excel)
        btn_layout.addWidget(export_excel_btn)

        btn_layout.addSpacing(20)

        export_osd_btn = QPushButton("📝 导出OSD")
        export_osd_btn.setToolTip("导出OSD名称到Excel表格，可编辑后导入")
        export_osd_btn.clicked.connect(self._export_osd)
        btn_layout.addWidget(export_osd_btn)

        import_osd_btn = QPushButton("📥 导入OSD")
        import_osd_btn.setToolTip("从Excel表格导入OSD名称并批量更新到设备")
        import_osd_btn.clicked.connect(self._import_osd)
        btn_layout.addWidget(import_osd_btn)

        btn_layout.addStretch()

        close_btn = QPushButton("❌ 关闭")
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)

    # ------------------------------------------------------------------ #
    #  导出 CSV
    # ------------------------------------------------------------------ #

    def _export_csv(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出CSV文件",
            f"{self.device_name}_通道信息.csv",
            "CSV文件 (*.csv)"
        )
        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                headers = [
                    "通道号", "通道名称", "在线状态", "IP地址", "协议", "OSD名称",
                    "主码流分辨率", "主码流编码", "主码流码率控制", "主码流码率(kbps)", "主码流帧率(fps)",
                    "子码流分辨率", "子码流编码", "子码流码率控制", "子码流码率(kbps)", "子码流帧率(fps)"
                ]
                writer.writerow(headers)
                for row in self.table_data:
                    writer.writerow([
                        row['channel_no'], row['channel_name'], row['online'],
                        row['ip'], row['protocol'], row['osd_name'],
                        row['main_resolution'], row['main_codec'], row['main_bitrate_mode'],
                        row['main_bitrate'], row['main_fps'],
                        row['sub_resolution'], row['sub_codec'], row['sub_bitrate_mode'],
                        row['sub_bitrate'], row['sub_fps'],
                    ])

            QMessageBox.information(self, "导出成功", f"已成功导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出CSV失败:\n{str(e)}")

    # ------------------------------------------------------------------ #
    #  导出 Excel
    # ------------------------------------------------------------------ #

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(
                self, "缺少依赖",
                "导出Excel需要安装openpyxl库\n请运行: pip install openpyxl\n\n或使用CSV导出功能。"
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel文件",
            f"{self.device_name}_通道信息.xlsx",
            "Excel文件 (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "通道信息"

            ws.merge_cells('A1:P1')
            title_cell = ws['A1']
            title_cell.value = f"{self.device_name} - 通道信息"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            headers = [
                "通道号", "通道名称", "在线状态", "IP地址", "协议", "OSD名称",
                "主码流分辨率", "主码流编码", "主码流码率控制", "主码流码率(kbps)", "主码流帧率(fps)",
                "子码流分辨率", "子码流编码", "子码流码率控制", "子码流码率(kbps)", "子码流帧率(fps)"
            ]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row_idx, row_data in enumerate(self.table_data, 3):
                ws.cell(row=row_idx, column=1, value=row_data['channel_no'])
                ws.cell(row=row_idx, column=2, value=row_data['channel_name'])
                ws.cell(row=row_idx, column=3, value=row_data['online'])
                ws.cell(row=row_idx, column=4, value=row_data['ip'])
                ws.cell(row=row_idx, column=5, value=row_data['protocol'])
                ws.cell(row=row_idx, column=6, value=row_data['osd_name'])
                ws.cell(row=row_idx, column=7, value=row_data['main_resolution'])
                ws.cell(row=row_idx, column=8, value=row_data['main_codec'])
                ws.cell(row=row_idx, column=9, value=row_data['main_bitrate_mode'])
                ws.cell(row=row_idx, column=10, value=int(row_data['main_bitrate'].replace(' kbps', '')) if row_data['main_bitrate'] != '-' else 0)
                ws.cell(row=row_idx, column=11, value=int(row_data['main_fps']) if row_data['main_fps'] != '-' else 0)
                ws.cell(row=row_idx, column=12, value=row_data['sub_resolution'])
                ws.cell(row=row_idx, column=13, value=row_data['sub_codec'])
                ws.cell(row=row_idx, column=14, value=row_data['sub_bitrate_mode'])
                ws.cell(row=row_idx, column=15, value=int(row_data['sub_bitrate'].replace(' kbps', '')) if row_data['sub_bitrate'] != '-' else 0)
                ws.cell(row=row_idx, column=16, value=int(row_data['sub_fps']) if row_data['sub_fps'] != '-' else 0)

                online_cell = ws.cell(row=row_idx, column=3)
                if row_data['online'] == '在线':
                    online_cell.font = Font(color="00B050")
                else:
                    online_cell.font = Font(color="FF0000")

            col_widths = [8, 15, 10, 15, 10, 15, 15, 25, 12, 15, 12, 15, 25, 12, 15, 12]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"已成功导出到:\n{file_path}")

        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出Excel失败:\n{str(e)}")

    # ------------------------------------------------------------------ #
    #  OSD 导出 / 导入
    # ------------------------------------------------------------------ #

    def _export_osd(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出OSD文件",
            f"{self.device_name}_OSD配置.xlsx",
            "Excel文件 (*.xlsx)"
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "OSD配置"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            headers = ["通道号", "当前OSD名称", "新OSD名称(在此列修改)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, row in enumerate(self.table_data, 2):
                osd_name = row.get('osd_name', '')
                ws.cell(row=row_idx, column=1, value=row['channel_no']).border = thin_border
                ws.cell(row=row_idx, column=2, value=osd_name).border = thin_border
                ws.cell(row=row_idx, column=3, value="").border = thin_border

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 35

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            wb.save(file_path)

            QMessageBox.information(self, "导出成功",
                f"OSD配置已导出到:\n{file_path}\n\n"
                f"使用说明：\n"
                f"1. 在'新OSD名称'列填写需要修改的通道名称\n"
                f"2. 不需要修改的通道请保持该列为空\n"
                f"3. 保存后使用'导入OSD'功能批量更新到设备\n\n"
                f"注意：只有填写了新名称的通道才会被更新，\n"
                f"      为空的通道将保持原名称不变。")
        except ImportError:
            QMessageBox.warning(self, "缺少依赖",
                "导出Excel需要安装openpyxl库\n请运行: pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出OSD失败:\n{str(e)}")

    def _import_osd(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "导入OSD文件", "",
            "Excel文件 (*.xlsx);;CSV文件 (*.csv)"
        )
        if not file_path:
            return

        try:
            osd_updates = []

            if file_path.lower().endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    headers = next(reader)
                    for row in reader:
                        if len(row) >= 3:
                            try:
                                ch_no = int(row[0])
                                new_osd = row[2].strip()
                                if new_osd:
                                    osd_updates.append((ch_no, new_osd))
                            except (ValueError, IndexError):
                                continue
            else:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 3:
                        try:
                            ch_no = int(row[0]) if row[0] else None
                            new_osd = str(row[2]).strip() if row[2] else ""
                            if ch_no and new_osd:
                                osd_updates.append((ch_no, new_osd))
                        except (ValueError, TypeError):
                            continue

            if not osd_updates:
                QMessageBox.information(self, "提示", "没有找到需要更新的OSD配置\n\n"
                    "导入规则：\n"
                    "• 只更新'新OSD名称'列有填写内容的通道\n"
                    "• 如果'新OSD名称'为空，则跳过该通道（保持原名称不变）")
                return

            update_details = "\n".join([f"  通道{ch}: {name}" for ch, name in osd_updates[:10]])
            if len(osd_updates) > 10:
                update_details += f"\n  ... 等共 {len(osd_updates)} 个通道"

            reply = QMessageBox.question(
                self, "确认更新",
                f"找到 {len(osd_updates)} 个通道的OSD配置需要更新\n\n"
                f"{update_details}\n\n"
                f"是否确认批量更新到设备 {self.device_name}?\n\n"
                f"注意：未填写新OSD名称的通道将保持原名称不变。",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply != QMessageBox.Yes:
                return

            self._start_osd_update(osd_updates)

        except ImportError:
            QMessageBox.warning(self, "缺少依赖",
                "导入Excel需要安装openpyxl库\n请运行: pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"导入OSD失败:\n{str(e)}")

    def _start_osd_update(self, osd_updates: list):
        if not self.device_config:
            QMessageBox.warning(self, "错误", "缺少设备配置信息，无法更新OSD")
            return

        self.osd_update_signal.emit(osd_updates, self.device_config)

        QMessageBox.information(self, "提示",
            f"OSD更新任务已提交，共 {len(osd_updates)} 个通道\n"
            f"请在主窗口查看更新进度。")
