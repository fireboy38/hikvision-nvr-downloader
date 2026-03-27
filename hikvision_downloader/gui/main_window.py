# 海康NVR批量录像下载工具 - 主窗口（SDK版）
import sys
import os
import json
import threading
import time
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

from PyQt5.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QPushButton, QLabel, QLineEdit,
    QSpinBox, QDateTimeEdit, QProgressBar, QTextEdit,
    QGroupBox, QFormLayout, QDialog, QDialogButtonBox, QMessageBox,
    QFileDialog, QStatusBar, QMenuBar, QMenu, QAction, QToolBar,
    QHeaderView, QAbstractItemView, QSplitter, QAbstractScrollArea,
    QListWidget, QListWidgetItem, QTreeWidget, QTreeWidgetItem,
    QApplication, QCheckBox, QComboBox,
)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QDateTime, QThread, QTime, QDate
from PyQt5.QtGui import QFont, QColor

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from core.nvr_api import HikvisionISAPI, create_isapi
from core.downloader import DownloadManager, DownloadTask, DownloadStatus, BatchDownloader


# ================================================================= #
#  颜色 / 状态工具
# ================================================================= #

STATUS_COLORS = {
    DownloadStatus.PENDING:     QColor(128, 128, 128),
    DownloadStatus.DOWNLOADING: QColor(0, 120, 215),
    DownloadStatus.MERGING:     QColor(255, 152, 0),   # 橙色
    DownloadStatus.COMPLETED:   QColor(0, 153, 76),
    DownloadStatus.FAILED:      QColor(232, 17, 35),
    DownloadStatus.CANCELLED:   QColor(160, 160, 160),
}
STATUS_TEXT = {
    DownloadStatus.PENDING:     "等待中",
    DownloadStatus.DOWNLOADING: "下载中",
    DownloadStatus.MERGING:     "合并中",
    DownloadStatus.COMPLETED:   "已完成",
    DownloadStatus.FAILED:      "失败",
    DownloadStatus.CANCELLED:   "已取消",
}


# ================================================================= #
#  设备配置对话框
# ================================================================= #

class DeviceConfigDialog(QDialog):
    def __init__(self, parent=None, device: Dict = None):
        super().__init__(parent)
        self.device = device or {}
        self.setWindowTitle("设备配置")
        self.setMinimumWidth(420)
        self._build_ui()

    def _build_ui(self):
        layout = QFormLayout(self)
        layout.setSpacing(10)

        self.name_edit = QLineEdit(self.device.get('name', ''))
        self.name_edit.setPlaceholderText("例：主楼NVR")
        layout.addRow("设备名称:", self.name_edit)

        self.host_edit = QLineEdit(self.device.get('host', ''))
        self.host_edit.setPlaceholderText("NVR IP地址")
        layout.addRow("IP地址:", self.host_edit)

        self.sdk_port = QSpinBox()
        self.sdk_port.setRange(1, 65535)
        self.sdk_port.setValue(self.device.get('port', 8000))
        layout.addRow("SDK端口:", self.sdk_port)

        self.http_port = QSpinBox()
        self.http_port.setRange(1, 65535)
        self.http_port.setValue(self.device.get('http_port', 80))
        layout.addRow("HTTP端口:", self.http_port)

        self.username_edit = QLineEdit(self.device.get('username', 'admin'))
        layout.addRow("用户名:", self.username_edit)

        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)
        if self.device.get('password'):
            self.password_edit.setText(self.device['password'])
        layout.addRow("密码:", self.password_edit)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addRow(btns)

    def get_config(self) -> Dict:
        host = self.host_edit.text().strip()
        return {
            'name':      self.name_edit.text().strip() or host,
            'host':      host,
            'port':      self.sdk_port.value(),
            'http_port': self.http_port.value(),
            'username':  self.username_edit.text().strip(),
            'password':  self.password_edit.text(),
        }


# ================================================================= #
#  时间预设管理对话框
# ================================================================= #

class TimePresetDialog(QDialog):
    """时间预设管理对话框"""
    def __init__(self, parent=None, presets: Dict = None):
        super().__init__(parent)
        self.setWindowTitle("管理时间预设")
        self.setMinimumWidth(500)
        self.setMinimumHeight(400)
        self.presets = presets or {}
        self._build_ui()
        self._load_presets()
    
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        # 说明标签
        hint = QLabel("自定义常用时间段，如：语文考试、早自习、课间操等")
        hint.setStyleSheet("color: #666; font-size: 12px;")
        layout.addWidget(hint)
        
        # 预设列表
        self._list = QListWidget()
        self._list.setSelectionMode(QAbstractItemView.SingleSelection)
        self._list.itemClicked.connect(self._on_item_selected)
        layout.addWidget(self._list)
        
        # 编辑区域
        edit_group = QGroupBox("编辑预设")
        edit_layout = QFormLayout()
        
        self._name_edit = QLineEdit()
        self._name_edit.setPlaceholderText("例如：语文考试")
        edit_layout.addRow("预设名称:", self._name_edit)
        
        self._start_edit = QDateTimeEdit()
        self._start_edit.setCalendarPopup(True)
        self._start_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self._start_edit.setDateTime(QDateTime.currentDateTime().addSecs(-3600))
        edit_layout.addRow("开始时间:", self._start_edit)
        
        self._end_edit = QDateTimeEdit()
        self._end_edit.setCalendarPopup(True)
        self._end_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self._end_edit.setDateTime(QDateTime.currentDateTime())
        edit_layout.addRow("结束时间:", self._end_edit)
        
        # 快捷设置今天/昨天
        quick_layout = QHBoxLayout()
        btn_today = QPushButton("设为今天")
        btn_today.clicked.connect(lambda: self._set_date_range("today"))
        quick_layout.addWidget(btn_today)
        
        btn_yesterday = QPushButton("设为昨天")
        btn_yesterday.clicked.connect(lambda: self._set_date_range("yesterday"))
        quick_layout.addWidget(btn_yesterday)
        
        btn_1h = QPushButton("1小时")
        btn_1h.clicked.connect(lambda: self._set_duration(3600))
        quick_layout.addWidget(btn_1h)
        
        btn_2h = QPushButton("2小时")
        btn_2h.clicked.connect(lambda: self._set_duration(7200))
        quick_layout.addWidget(btn_2h)
        
        edit_layout.addRow("快捷设置:", quick_layout)
        
        edit_group.setLayout(edit_layout)
        layout.addWidget(edit_group)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        self._btn_add = QPushButton("➕ 添加")
        self._btn_add.clicked.connect(self._add_preset)
        btn_layout.addWidget(self._btn_add)
        
        self._btn_update = QPushButton("💾 保存修改")
        self._btn_update.clicked.connect(self._update_preset)
        btn_layout.addWidget(self._btn_update)
        
        self._btn_delete = QPushButton("🗑️ 删除")
        self._btn_delete.clicked.connect(self._delete_preset)
        btn_layout.addWidget(self._btn_delete)
        
        btn_layout.addStretch()
        
        self._btn_clear = QPushButton("清空")
        self._btn_clear.clicked.connect(self._clear_form)
        btn_layout.addWidget(self._btn_clear)
        
        layout.addLayout(btn_layout)
        
        # 确定/取消按钮
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)
    
    def _set_date_range(self, preset: str):
        """设置日期范围"""
        now = QDateTime.currentDateTime()
        if preset == "today":
            self._start_edit.setDateTime(QDateTime(now.date(), QTime(0, 0, 0)))
            self._end_edit.setDateTime(QDateTime(now.date(), QTime(23, 59, 59)))
        elif preset == "yesterday":
            yd = now.date().addDays(-1)
            self._start_edit.setDateTime(QDateTime(yd, QTime(0, 0, 0)))
            self._end_edit.setDateTime(QDateTime(yd, QTime(23, 59, 59)))
    
    def _set_duration(self, seconds: int):
        """设置持续时间"""
        start = self._start_edit.dateTime()
        self._end_edit.setDateTime(start.addSecs(seconds))
    
    def _load_presets(self):
        """加载预设到列表"""
        self._list.clear()
        for name, data in self.presets.items():
            item = QListWidgetItem(f"{name} ({data['start']} ~ {data['end']})")
            item.setData(Qt.UserRole, name)
            self._list.addItem(item)
    
    def _on_item_selected(self, item):
        """选中列表项时填充表单"""
        name = item.data(Qt.UserRole)
        if name in self.presets:
            data = self.presets[name]
            self._name_edit.setText(name)
            self._start_edit.setDateTime(QDateTime.fromString(data['start'], "yyyy-MM-dd HH:mm:ss"))
            self._end_edit.setDateTime(QDateTime.fromString(data['end'], "yyyy-MM-dd HH:mm:ss"))
    
    def _add_preset(self):
        """添加新预设"""
        name = self._name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "提示", "请输入预设名称")
            return
        
        if name in self.presets:
            QMessageBox.warning(self, "提示", f"预设 '{name}' 已存在，请使用保存修改")
            return
        
        self.presets[name] = {
            'start': self._start_edit.dateTime().toString("yyyy-MM-dd HH:mm:ss"),
            'end': self._end_edit.dateTime().toString("yyyy-MM-dd HH:mm:ss")
        }
        self._load_presets()
        QMessageBox.information(self, "成功", f"已添加预设 '{name}'")
    
    def _update_preset(self):
        """更新现有预设"""
        name = self._name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "提示", "请输入预设名称")
            return
        
        if name not in self.presets:
            QMessageBox.warning(self, "提示", f"预设 '{name}' 不存在，请使用添加")
            return
        
        self.presets[name] = {
            'start': self._start_edit.dateTime().toString("yyyy-MM-dd HH:mm:ss"),
            'end': self._end_edit.dateTime().toString("yyyy-MM-dd HH:mm:ss")
        }
        self._load_presets()
        QMessageBox.information(self, "成功", f"已更新预设 '{name}'")
    
    def _delete_preset(self):
        """删除预设"""
        name = self._name_edit.text().strip()
        if not name or name not in self.presets:
            QMessageBox.warning(self, "提示", "请先选择要删除的预设")
            return
        
        reply = QMessageBox.question(self, "确认", f"确定要删除预设 '{name}' 吗？",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            del self.presets[name]
            self._load_presets()
            self._clear_form()
    
    def _clear_form(self):
        """清空表单"""
        self._name_edit.clear()
        self._start_edit.setDateTime(QDateTime.currentDateTime().addSecs(-3600))
        self._end_edit.setDateTime(QDateTime.currentDateTime())
        self._list.clearSelection()
    
    def get_presets(self) -> Dict:
        """返回所有预设"""
        return self.presets


# ================================================================= #
#  设置对话框
# ================================================================= #

class DownloadSettingsDialog(QDialog):
    """下载设置对话框"""
    def __init__(self, parent=None,
                 total_thread_count=9, per_device_thread_count=3,
                 merge_mode='standard', enable_debug_log=True, skip_transcode=True,
                 download_mode='isapi'):
        super().__init__(parent)
        self.setWindowTitle("下载设置")
        self.setMinimumWidth(450)
        self._build_ui()

        self._total_thread_spin.setValue(total_thread_count)
        self._per_device_thread_spin.setValue(per_device_thread_count)
        # 设置合并模式下拉框的当前选中项
        merge_mode_index = {'ultra': 0, 'fast': 1, 'standard': 2}.get(merge_mode, 0)
        self._merge_mode_combo.setCurrentIndex(merge_mode_index)
        self._debug_log_chk.setChecked(enable_debug_log)
        self._skip_transcode_chk.setChecked(skip_transcode)
        # 设置下载模式
        mode_index = {'isapi': 0, 'sdk': 1}.get(download_mode, 0)
        self._download_mode_combo.setCurrentIndex(mode_index)


    def _build_ui(self):
        layout = QFormLayout(self)
        layout.setSpacing(12)

        # ===== 下载模式 =====
        mode_group = QLabel("<b>下载模式</b>")
        layout.addRow(mode_group)

        self._download_mode_combo = QComboBox()
        self._download_mode_combo.addItem("🌐 ISAPI模式（推荐）", "isapi")
        self._download_mode_combo.addItem("🔌 SDK模式", "sdk")
        # self._download_mode_combo.addItem("📦 HikLoad模式", "hikload")  # HikLoad模式已禁用
        self._download_mode_combo.setCurrentIndex(0)  # 默认ISAPI
        layout.addRow("下载模式:", self._download_mode_combo)

        mode_hint = QLabel("<small>ISAPI模式：纯HTTP下载，速度快且稳定，无需Java SDK<br/>SDK模式：通过海康HCNetSDK.dll下载，支持分段+合并转码</small>")
        mode_hint.setWordWrap(True)
        mode_hint.setStyleSheet("color: #666;")
        layout.addRow("", mode_hint)

        layout.addRow(QLabel(""))

        # ===== 线程设置 =====
        thread_group = QLabel("<b>线程设置</b>")
        layout.addRow(thread_group)
        
        # 总下载线程数
        self._total_thread_spin = QSpinBox()
        self._total_thread_spin.setRange(1, 20)
        self._total_thread_spin.setValue(9)  # 默认9线程
        self._total_thread_spin.setSuffix(" 线程")
        layout.addRow("总下载线程数:", self._total_thread_spin)
        
        total_thread_hint = QLabel("<small>全局下载线程池大小，建议：9线程（1-20）<br/>多台NVR时可提高此值</small>")
        total_thread_hint.setWordWrap(True)
        total_thread_hint.setStyleSheet("color: #666;")
        layout.addRow("", total_thread_hint)
        
        # 每台NVR线程数
        self._per_device_thread_spin = QSpinBox()
        self._per_device_thread_spin.setRange(1, 6)
        self._per_device_thread_spin.setValue(3)  # 默认3线程
        self._per_device_thread_spin.setSuffix(" 线程")
        layout.addRow("每台NVR并发数:", self._per_device_thread_spin)
        
        per_device_hint = QLabel("<small>单台NVR最大并发连接数，建议：3线程（1-6）<br/>海康NVR每用户最多4个SDK连接</small>")
        per_device_hint.setWordWrap(True)
        per_device_hint.setStyleSheet("color: #666;")
        layout.addRow("", per_device_hint)

        layout.addRow(QLabel(""))  # 分隔

        # 合并模式设置
        self._merge_mode_combo = QComboBox()
        self._merge_mode_combo.addItem("极速模式（不转码，无faststart）", "ultra")
        self._merge_mode_combo.addItem("快速模式（不转码，有faststart）", "fast")
        self._merge_mode_combo.addItem("标准模式（转码合并，兼容性好）", "standard")
        self._merge_mode_combo.setCurrentIndex(0)  # 默认极速模式
        layout.addRow("合并模式:", self._merge_mode_combo)

        merge_hint = QLabel("<small>极速模式：最快，适合本地播放<br/>快速模式：较快，适合网络播放<br/>标准模式：较慢，兼容性最好</small>")
        merge_hint.setWordWrap(True)
        merge_hint.setStyleSheet("color: #666;")
        layout.addRow("", merge_hint)

        layout.addRow(QLabel(""))  # 分隔

        # 调试日志设置
        self._debug_log_chk = QCheckBox("启用调试日志")
        self._debug_log_chk.setChecked(True)  # 默认开启
        self._debug_log_chk.setToolTip("生成详细的下载和合并日志，用于排查合并点问题")
        layout.addRow("", self._debug_log_chk)

        debug_hint = QLabel("<small>启用后会在下载目录生成详细日志文件<br/>包含分段信息、合并点时间戳等</small>")
        debug_hint.setWordWrap(True)
        debug_hint.setStyleSheet("color: #666;")
        layout.addRow("", debug_hint)

        layout.addRow(QLabel(""))  # 分隔

        # 转码设置
        self._skip_transcode_chk = QCheckBox("跳过转码（原始格式）")
        self._skip_transcode_chk.setChecked(True)  # 默认跳过
        self._skip_transcode_chk.setToolTip("跳过FFmpeg转码，直接使用NVR原始下载的文件。速度更快，原始文件通常可正常播放。")
        layout.addRow("", self._skip_transcode_chk)

        transcode_hint = QLabel("<small>推荐：勾选跳过转码（速度更快）<br/>如果播放问题可取消勾选，转换为标准MP4</small>")
        transcode_hint.setWordWrap(True)
        transcode_hint.setStyleSheet("color: #666;")
        layout.addRow("", transcode_hint)

        # 按钮
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addRow(btns)

    def get_settings(self):
        return {
            'total_thread_count': self._total_thread_spin.value(),
            'per_device_thread_count': self._per_device_thread_spin.value(),
            'merge_mode': self._merge_mode_combo.currentData(),
            'enable_debug_log': self._debug_log_chk.isChecked(),
            'skip_transcode': self._skip_transcode_chk.isChecked(),
            'download_mode': self._download_mode_combo.currentData(),
        }



class ChannelInfoDialog(QDialog):
    """通道信息表格展示对话框，支持导出Excel/CSV"""
    
    # OSD更新信号
    osd_update_signal = pyqtSignal(list, dict)  # osd_updates, device_config
    
    def __init__(self, device_name: str, table_data: List[Dict], device_config: dict = None, parent=None):
        super().__init__(parent)
        self.device_name = device_name
        self.table_data = table_data
        self.device_config = device_config or {}
        self.setWindowTitle(f"通道信息 - {device_name}")
        self.setMinimumSize(1400, 600)
        self.resize(1600, 750)
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        # 标题
        title_label = QLabel(f"📹 {self.device_name} - 通道信息 ({len(self.table_data)}个通道)")
        title_font = QFont()
        title_font.setPointSize(12)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 统计信息
        online_count = sum(1 for row in self.table_data if row['online'] == '在线')
        stats_label = QLabel(f"在线: {online_count} | 离线: {len(self.table_data) - online_count}")
        stats_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(stats_label)
        
        # 表格
        self.table = QTableWidget()
        self.table.setColumnCount(16)
        headers = [
            "通道号", "通道名称", "在线状态", "IP地址", "协议", "OSD名称",
            "主码流分辨率", "主码流编码", "主码流码率控制", "主码流码率", "主码流帧率",
            "子码流分辨率", "子码流编码", "子码流码率控制", "子码流码率", "子码流帧率"
        ]
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.table_data))
        
        # 设置列宽 - 根据内容自适应
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(False)
        
        # 设置最小列宽确保可读性
        min_widths = [50, 100, 60, 100, 70, 100, 90, 140, 90, 70, 60, 90, 140, 90, 70, 60]
        for i, width in enumerate(min_widths):
            self.table.setColumnWidth(i, width)
        
        # 填充数据
        for row_idx, row_data in enumerate(self.table_data):
            # 通道号
            item = QTableWidgetItem(str(row_data['channel_no']))
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 0, item)
            
            # 通道名称
            self.table.setItem(row_idx, 1, QTableWidgetItem(row_data['channel_name']))
            
            # 在线状态（带颜色）
            online_item = QTableWidgetItem(row_data['online'])
            online_item.setTextAlignment(Qt.AlignCenter)
            if row_data['online'] == '在线':
                online_item.setForeground(QColor(0, 153, 76))  # 绿色
            else:
                online_item.setForeground(QColor(232, 17, 35))  # 红色
            self.table.setItem(row_idx, 2, online_item)
            
            # IP地址
            self.table.setItem(row_idx, 3, QTableWidgetItem(row_data['ip']))
            
            # 协议
            self.table.setItem(row_idx, 4, QTableWidgetItem(row_data['protocol']))
            
            # OSD名称
            self.table.setItem(row_idx, 5, QTableWidgetItem(row_data['osd_name']))
            
            # 主码流信息
            self.table.setItem(row_idx, 6, QTableWidgetItem(row_data['main_resolution']))
            self.table.setItem(row_idx, 7, QTableWidgetItem(row_data['main_codec']))
            
            main_bitrate_mode_item = QTableWidgetItem(row_data['main_bitrate_mode'])
            main_bitrate_mode_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 8, main_bitrate_mode_item)
            
            main_bitrate_item = QTableWidgetItem(row_data['main_bitrate'])
            main_bitrate_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 9, main_bitrate_item)
            
            main_fps_item = QTableWidgetItem(row_data['main_fps'])
            main_fps_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 10, main_fps_item)
            
            # 子码流信息
            self.table.setItem(row_idx, 11, QTableWidgetItem(row_data['sub_resolution']))
            self.table.setItem(row_idx, 12, QTableWidgetItem(row_data['sub_codec']))
            
            sub_bitrate_mode_item = QTableWidgetItem(row_data['sub_bitrate_mode'])
            sub_bitrate_mode_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 13, sub_bitrate_mode_item)
            
            sub_bitrate_item = QTableWidgetItem(row_data['sub_bitrate'])
            sub_bitrate_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 14, sub_bitrate_item)
            
            sub_fps_item = QTableWidgetItem(row_data['sub_fps'])
            sub_fps_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 15, sub_fps_item)
        
        # 设置表格属性
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # 禁用水平滚动条，调整表格大小以适应内容
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        
        # 根据内容调整表格大小
        self.table.resizeColumnsToContents()
        
        layout.addWidget(self.table)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        # 导出CSV按钮
        export_csv_btn = QPushButton("📄 导出CSV")
        export_csv_btn.setToolTip("导出为CSV格式表格文件")
        export_csv_btn.clicked.connect(self._export_csv)
        btn_layout.addWidget(export_csv_btn)
        
        # 导出Excel按钮
        export_excel_btn = QPushButton("📊 导出Excel")
        export_excel_btn.setToolTip("导出为Excel格式表格文件（需要安装openpyxl）")
        export_excel_btn.clicked.connect(self._export_excel)
        btn_layout.addWidget(export_excel_btn)
        
        btn_layout.addSpacing(20)
        
        # OSD批量操作按钮
        export_osd_btn = QPushButton("📝 导出OSD")
        export_osd_btn.setToolTip("导出OSD名称到Excel表格，可编辑后导入")
        export_osd_btn.clicked.connect(self._export_osd)
        btn_layout.addWidget(export_osd_btn)
        
        import_osd_btn = QPushButton("📥 导入OSD")
        import_osd_btn.setToolTip("从Excel表格导入OSD名称并批量更新到设备")
        import_osd_btn.clicked.connect(self._import_osd)
        btn_layout.addWidget(import_osd_btn)
        
        btn_layout.addStretch()
        
        # 关闭按钮
        close_btn = QPushButton("❌ 关闭")
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
    
    def _export_csv(self):
        """导出为CSV文件"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出CSV文件", 
            f"{self.device_name}_通道信息.csv",
            "CSV文件 (*.csv)"
        )
        if not file_path:
            return
        
        try:
            import csv
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                # 写入表头
                headers = [
                    "通道号", "通道名称", "在线状态", "IP地址", "协议", "OSD名称",
                    "主码流分辨率", "主码流编码", "主码流码率控制", "主码流码率(kbps)", "主码流帧率(fps)",
                    "子码流分辨率", "子码流编码", "子码流码率控制", "子码流码率(kbps)", "子码流帧率(fps)"
                ]
                writer.writerow(headers)
                # 写入数据
                for row in self.table_data:
                    writer.writerow([
                        row['channel_no'],
                        row['channel_name'],
                        row['online'],
                        row['ip'],
                        row['protocol'],
                        row['osd_name'],
                        row['main_resolution'],
                        row['main_codec'],
                        row['main_bitrate_mode'],
                        row['main_bitrate'],
                        row['main_fps'],
                        row['sub_resolution'],
                        row['sub_codec'],
                        row['sub_bitrate_mode'],
                        row['sub_bitrate'],
                        row['sub_fps'],
                    ])
            
            QMessageBox.information(self, "导出成功", f"已成功导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出CSV失败:\n{str(e)}")
    
    def _export_excel(self):
        """导出为Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(
                self, "缺少依赖", 
                "导出Excel需要安装openpyxl库\n"
                "请运行: pip install openpyxl\n\n"
                "或使用CSV导出功能。"
            )
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel文件",
            f"{self.device_name}_通道信息.xlsx",
            "Excel文件 (*.xlsx)"
        )
        if not file_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "通道信息"
            
            # 设置标题
            ws.merge_cells('A1:P1')
            title_cell = ws['A1']
            title_cell.value = f"{self.device_name} - 通道信息"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 表头
            headers = [
                "通道号", "通道名称", "在线状态", "IP地址", "协议", "OSD名称",
                "主码流分辨率", "主码流编码", "主码流码率控制", "主码流码率(kbps)", "主码流帧率(fps)",
                "子码流分辨率", "子码流编码", "子码流码率控制", "子码流码率(kbps)", "子码流帧率(fps)"
            ]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数据
            for row_idx, row_data in enumerate(self.table_data, 3):
                ws.cell(row=row_idx, column=1, value=row_data['channel_no'])
                ws.cell(row=row_idx, column=2, value=row_data['channel_name'])
                ws.cell(row=row_idx, column=3, value=row_data['online'])
                ws.cell(row=row_idx, column=4, value=row_data['ip'])
                ws.cell(row=row_idx, column=5, value=row_data['protocol'])
                ws.cell(row=row_idx, column=6, value=row_data['osd_name'])
                ws.cell(row=row_idx, column=7, value=row_data['main_resolution'])
                ws.cell(row=row_idx, column=8, value=row_data['main_codec'])
                ws.cell(row=row_idx, column=9, value=row_data['main_bitrate_mode'])
                ws.cell(row=row_idx, column=10, value=int(row_data['main_bitrate'].replace(' kbps', '')) if row_data['main_bitrate'] != '-' else 0)
                ws.cell(row=row_idx, column=11, value=int(row_data['main_fps']) if row_data['main_fps'] != '-' else 0)
                ws.cell(row=row_idx, column=12, value=row_data['sub_resolution'])
                ws.cell(row=row_idx, column=13, value=row_data['sub_codec'])
                ws.cell(row=row_idx, column=14, value=row_data['sub_bitrate_mode'])
                ws.cell(row=row_idx, column=15, value=int(row_data['sub_bitrate'].replace(' kbps', '')) if row_data['sub_bitrate'] != '-' else 0)
                ws.cell(row=row_idx, column=16, value=int(row_data['sub_fps']) if row_data['sub_fps'] != '-' else 0)
                
                # 在线状态颜色
                online_cell = ws.cell(row=row_idx, column=3)
                if row_data['online'] == '在线':
                    online_cell.font = Font(color="00B050")  # 绿色
                else:
                    online_cell.font = Font(color="FF0000")  # 红色
            
            # 调整列宽
            col_widths = [8, 15, 10, 15, 10, 15, 15, 25, 12, 15, 12, 15, 25, 12, 15, 12]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"已成功导出到:\n{file_path}")
            
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出Excel失败:\n{str(e)}")
    
    def _export_osd(self):
        """导出OSD名称到Excel文件（简化格式，便于编辑）"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出OSD文件", 
            f"{self.device_name}_OSD配置.xlsx",
            "Excel文件 (*.xlsx)"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "OSD配置"
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            headers = ["通道号", "当前OSD名称", "新OSD名称(在此列修改)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 写入数据
            for row_idx, row in enumerate(self.table_data, 2):
                osd_name = row.get('osd_name', '')
                ws.cell(row=row_idx, column=1, value=row['channel_no']).border = thin_border
                ws.cell(row=row_idx, column=2, value=osd_name).border = thin_border
                ws.cell(row=row_idx, column=3, value="").border = thin_border  # 空列供用户填写新名称
            
            # 设置列宽
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 35
            
            # 设置数据行居中对齐
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(file_path)
            
            QMessageBox.information(self, "导出成功", 
                f"OSD配置已导出到:\n{file_path}\n\n"
                f"使用说明：\n"
                f"1. 在'新OSD名称'列填写需要修改的通道名称\n"
                f"2. 不需要修改的通道请保持该列为空\n"
                f"3. 保存后使用'导入OSD'功能批量更新到设备\n\n"
                f"注意：只有填写了新名称的通道才会被更新，\n"
                f"      为空的通道将保持原名称不变。")
        except ImportError:
            QMessageBox.warning(self, "缺少依赖", 
                "导出Excel需要安装openpyxl库\n"
                "请运行: pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出OSD失败:\n{str(e)}")
    
    def _import_osd(self):
        """从Excel文件导入OSD名称并批量更新"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "导入OSD文件",
            "",
            "Excel文件 (*.xlsx);;CSV文件 (*.csv)"
        )
        if not file_path:
            return
        
        try:
            osd_updates = []
            
            if file_path.lower().endswith('.csv'):
                # 兼容旧的CSV格式
                import csv
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    headers = next(reader)  # 跳过表头
                    
                    for row in reader:
                        if len(row) >= 3:
                            try:
                                ch_no = int(row[0])
                                new_osd = row[2].strip()  # 第3列是新OSD名称
                                if new_osd:  # 只处理填写了新名称的行
                                    osd_updates.append((ch_no, new_osd))
                            except (ValueError, IndexError):
                                continue
            else:
                # Excel格式
                from openpyxl import load_workbook
                
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 跳过表头，从第2行开始读取
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 3:
                        try:
                            ch_no = int(row[0]) if row[0] else None
                            new_osd = str(row[2]).strip() if row[2] else ""
                            if ch_no and new_osd:  # 只处理填写了新名称的行
                                osd_updates.append((ch_no, new_osd))
                        except (ValueError, TypeError):
                            continue
            
            if not osd_updates:
                QMessageBox.information(self, "提示", "没有找到需要更新的OSD配置\n\n"
                    "导入规则：\n"
                    "• 只更新'新OSD名称'列有填写内容的通道\n"
                    "• 如果'新OSD名称'为空，则跳过该通道（保持原名称不变）")
                return
            
            # 显示将要更新的通道列表
            update_details = "\n".join([f"  通道{ch}: {name}" for ch, name in osd_updates[:10]])
            if len(osd_updates) > 10:
                update_details += f"\n  ... 等共 {len(osd_updates)} 个通道"
            
            # 确认对话框
            reply = QMessageBox.question(
                self, "确认更新",
                f"找到 {len(osd_updates)} 个通道的OSD配置需要更新\n\n"
                f"{update_details}\n\n"
                f"是否确认批量更新到设备 {self.device_name}?\n\n"
                f"注意：未填写新OSD名称的通道将保持原名称不变。",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # 发送信号到主窗口进行批量更新
            self._start_osd_update(osd_updates)
            
        except ImportError:
            QMessageBox.warning(self, "缺少依赖", 
                "导入Excel需要安装openpyxl库\n"
                "请运行: pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"导入OSD失败:\n{str(e)}")
    
    def _start_osd_update(self, osd_updates: list):
        """开始批量更新OSD"""
        if not self.device_config:
            QMessageBox.warning(self, "错误", "缺少设备配置信息，无法更新OSD")
            return
        
        # 发送信号到主窗口进行批量更新
        self.osd_update_signal.emit(osd_updates, self.device_config)
        
        QMessageBox.information(self, "提示", 
            f"OSD更新任务已提交，共 {len(osd_updates)} 个通道\n"
            f"请在主窗口查看更新进度。")



# ================================================================= #
#  后台连接线程
# ================================================================= #

class ConnectWorker(QThread):
    """后台连接设备、获取通道的线程"""
    result_ready = pyqtSignal(bool, str, dict, list)   # ok, msg, dev_info, channels

    def __init__(self, config: Dict):
        super().__init__()
        self.config = config

    def run(self):
        try:
            from core.hcnetsdk import HCNetSDK
            cfg = self.config

            sdk = HCNetSDK()
            if not sdk.init():
                self.result_ready.emit(False, "SDK初始化失败", {}, [])
                return

            ok, msg, dev = sdk.login(
                cfg['host'], cfg.get('port', 8000),
                cfg.get('username', 'admin'), cfg.get('password', '')
            )
            if not ok:
                sdk.cleanup()
                self.result_ready.emit(False, f"登录失败: {msg}", {}, [])
                return

            # 用ISAPI补充通道名称
            channels = sdk.get_channels_with_names(
                total_ch   = dev['total_ch'],
                start_chan = max(dev.get('start_dchan', 1), 1),
                nvr_ip     = cfg['host'],
                nvr_port   = cfg.get('http_port', 80),
                username   = cfg.get('username', 'admin'),
                password   = cfg.get('password', ''),
            )

            sdk.cleanup()
            self.result_ready.emit(True, "连接成功", dev, channels)

        except Exception as e:
            self.result_ready.emit(False, f"异常: {e}", {}, [])


# ================================================================= #
#  主窗口
# ================================================================= #

class MainWindow(QMainWindow):
    _progress_signal = pyqtSignal(str, int)   # task_id, progress
    _status_signal   = pyqtSignal(str)         # task_id
    _log_signal = pyqtSignal(str)              # log message
    _size_signal = pyqtSignal(str, int)        # task_id, size_bytes（连接成功后立即更新录像大小）
    _estimate_signal = pyqtSignal(int, int, int)  # total_size, ok_count, fail_count（探测完成更新预计大小）
    _multi_connect_result_signal = pyqtSignal(dict, bool, str, dict, list)  # cfg, ok, msg, dev, channels
    _show_channel_info_signal = pyqtSignal(str, list)  # device_name, table_data


    def __init__(self):
        super().__init__()
        self.setWindowTitle("四川新数录像批量下载器")
        self.setMinimumSize(1280, 850)

        self.devices:      List[Dict] = []
        self._device_channels: Dict[str, List[Dict]] = {}  # {device_key: [channels]}
        self.download_dir: str = os.path.expanduser("~/Downloads")

        self._current_config: Optional[Dict] = None
        self._connect_worker: Optional[ConnectWorker] = None
        # 线程设置（总线程数和每台NVR线程数分离）
        self._total_thread_count = 9  # 默认总下载线程数
        self._per_device_thread_count = 3  # 默认每台NVR并发数
        
        self._merge_mode = "fast"  # 默认快速合并模式（不转码，有faststart）
        self._enable_debug_log = True  # 默认开启调试日志
        self._skip_transcode = True  # 默认跳过转码
        self._download_mode = "isapi"  # 下载模式: "isapi"（优先） 或 "sdk"
        
        # 初始化下载管理器（总线程9，每台NVR限制3，转码线程2）
        self._dm = DownloadManager(
            max_concurrent=self._total_thread_count,
            max_concurrent_per_device=self._per_device_thread_count,
            transcode_workers=2
        )
        self._batch: Optional[BatchDownloader] = None
        
        # 日志缓冲区
        self._run_log_buffer: List[str] = []
        self._download_log_buffer: List[str] = []
        self._log_view_mode = "run"  # 默认显示运行日志
        
        # 时间预设
        self._time_presets: Dict[str, Dict] = {}  # {名称: {start, end}}

        # 待下载任务列表（加入列表后等待下载的任务）
        self._pending_tasks: List[DownloadTask] = []

        # ISAPI模式停止事件集合
        self._isapi_stop_events: Dict[str, 'threading.Event'] = {}


        # 下载速度跟踪
        self._download_start_times: Dict[str, float] = {}  # {task_id: timestamp}
        self._task_file_sizes: Dict[str, int] = {}   # {task_id: bytes} (录像大小)

        # 网卡速度监控
        self._last_net_io = None
        self._last_net_time = 0

        self._load_config()
        self._build_ui()
        self._wire_signals()

    # ------------------------------------------------------------------ #
    #  UI构建
    # ------------------------------------------------------------------ #

    def _build_ui(self):
        self._make_menu()
        self._make_toolbar()

        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(6)

        # 左右面板
        hbox = QHBoxLayout()
        hbox.setSpacing(6)
        hbox.addWidget(self._make_left_panel(), 1)
        hbox.addWidget(self._make_right_panel(), 3)
        main_layout.addLayout(hbox)

        # 版权信息（在底部，状态栏上方）
        copyright_label = QLabel("版权所有：四川新数信息技术有限公司   www.scxs.vip")
        copyright_label.setStyleSheet("font-size: 11px; color: #888; padding: 5px;")
        copyright_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(copyright_label)

        self.statusBar().showMessage("就绪")

    def _make_menu(self):
        mb = self.menuBar()

        fm = mb.addMenu("文件")
        for label, slot in [
            ("添加设备", self._add_device),
            ("导入配置", self._import_config),
            ("导出配置", self._export_config),
        ]:
            a = QAction(label, self)
            a.triggered.connect(slot)
            fm.addAction(a)
        fm.addSeparator()
        
        # 设备列表导入导出
        a = QAction("📥 导入设备列表 (CSV)", self)
        a.triggered.connect(self._import_device_list)
        fm.addAction(a)
        a = QAction("📤 导出设备列表 (CSV)", self)
        a.triggered.connect(self._export_device_list)
        fm.addAction(a)
        fm.addSeparator()
        
        ex = QAction("退出", self)
        ex.triggered.connect(self.close)
        fm.addAction(ex)

        # 通道详情
        a = QAction("📋 通道详情", self)
        a.triggered.connect(lambda: self._on_query_channel_info_clicked())
        mb.addAction(a)

        # 设置菜单
        sm = mb.addMenu("设置")
        a = QAction("⚙️ 下载设置", self)
        a.triggered.connect(self._show_download_settings)
        sm.addAction(a)
        sm.addSeparator()
        a = QAction("⏰ 时间预设管理", self)
        a.triggered.connect(self._manage_time_presets)
        sm.addAction(a)

        # 关于菜单
        am = mb.addMenu("关于")
        a = QAction("📖 使用说明", self)
        a.triggered.connect(self._show_help)
        am.addAction(a)
        a = QAction("ℹ️ 关于软件", self)
        a.triggered.connect(self._show_about)
        am.addAction(a)

    def _make_toolbar(self):
        tb = QToolBar()
        tb.setMovable(False)
        self.addToolBar(tb)

        btns = [
            ("➕ 添加设备",  self._add_device),
            ("📋 通道详情",  self._query_channel_info),
            ("🔄 刷新通道",  self._refresh_channels),
        ]
        for label, slot in btns:
            b = QPushButton(label)
            b.clicked.connect(slot)
            tb.addWidget(b)

        tb.addSeparator()

        self._btn_add_tasks = QPushButton("📋 加入列表")
        self._btn_add_tasks.clicked.connect(self._add_tasks_to_list)
        self._btn_add_tasks.setEnabled(False)
        self._btn_add_tasks.setStyleSheet("QPushButton { background-color: #2196F3; color: white; font-weight: bold; padding: 5px 12px; }")
        tb.addWidget(self._btn_add_tasks)

        tb.addSeparator()

        self._btn_settings = QPushButton("⚙️ 设置")
        self._btn_settings.clicked.connect(self._show_download_settings)
        tb.addWidget(self._btn_settings)

        # 下载模式切换（ISAPI优先）
        tb.addSeparator()
        mode_label = QLabel("  下载模式:")
        mode_label.setStyleSheet("color: #555; font-weight: bold;")
        tb.addWidget(mode_label)

        self._mode_combo = QComboBox()
        self._mode_combo.addItem("🌐 ISAPI模式（推荐）", "isapi")
        self._mode_combo.addItem("🔌 SDK模式", "sdk")
        # self._mode_combo.addItem("📦 HikLoad模式", "hikload")  # HikLoad模式已禁用
        self._mode_combo.setFixedWidth(180)
        # 根据当前模式设置选中项
        idx = 0 if self._download_mode == "isapi" else 1
        self._mode_combo.setCurrentIndex(idx)
        self._mode_combo.currentIndexChanged.connect(self._on_download_mode_changed)
        self._mode_combo.setToolTip("ISAPI模式：纯HTTP下载，无需Java SDK，速度快且稳定\nSDK模式：通过海康HCNetSDK.dll下载，支持分段+合并")
        tb.addWidget(self._mode_combo)

        tb.addSeparator()

        self._btn_dir = QPushButton("📁 保存目录")
        self._btn_dir.clicked.connect(self._pick_dir)
        tb.addWidget(self._btn_dir)

        self._dir_label = QLabel(f"  {self.download_dir}")
        self._dir_label.setStyleSheet("color: #555;")
        tb.addWidget(self._dir_label)

    def _make_left_panel(self) -> QWidget:
        w = QWidget()
        vbox = QVBoxLayout(w)
        vbox.setSpacing(6)

        # 设备管理
        dg = QGroupBox("设备管理")
        dl = QVBoxLayout()

        self._device_list = QListWidget()
        self._device_list.setSelectionMode(QAbstractItemView.ExtendedSelection)  # 支持多选
        self._device_list.itemDoubleClicked.connect(self._on_device_double_clicked)  # 双击连接设备
        self._device_list.currentRowChanged.connect(self._on_device_row_changed)
        dl.addWidget(self._device_list)

        hb = QHBoxLayout()
        for label, slot in [("添加", self._add_device), ("编辑", self._edit_device),
                             ("删除", self._del_device), ("查询", self._query_device)]:
            b = QPushButton(label)
            b.clicked.connect(slot)
            hb.addWidget(b)
        dl.addLayout(hb)
        dg.setLayout(dl)
        vbox.addWidget(dg, 1)  # 设备管理窗口垂直比例小一些

        # 通道选择
        cg = QGroupBox("通道选择")
        cl = QVBoxLayout()

        self._channel_tree = QTreeWidget()
        self._channel_tree.setHeaderLabels(["设备/通道", ""])
        self._channel_tree.setColumnWidth(0, 200)
        self._channel_tree.setColumnHidden(1, True)  # 隐藏第2列
        self._channel_tree.itemChanged.connect(self._update_main_channel_count)
        cl.addWidget(self._channel_tree)

        hb2 = QHBoxLayout()
        sel_all = QPushButton("全选")
        sel_all.clicked.connect(self._select_all)
        hb2.addWidget(sel_all)
        desel = QPushButton("取消全选")
        desel.clicked.connect(self._deselect_all)
        hb2.addWidget(desel)
        
        btn_add = QPushButton("📋 加入列表")
        btn_add.clicked.connect(self._add_tasks_to_list)
        btn_add.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }")
        hb2.addWidget(btn_add)
        
        cl.addLayout(hb2)

        self._channel_count_label = QLabel("共 0 个通道，已选择 0 个")
        self._channel_count_label.setStyleSheet("color:#666;font-size:11px;")
        cl.addWidget(self._channel_count_label)

        cg.setLayout(cl)
        vbox.addWidget(cg, 3)  # 通道选择窗口垂直比例大一些

        # 时间选择
        tg = QGroupBox("录像时间范围")
        tl = QFormLayout()

        self._dt_start = QDateTimeEdit()
        self._dt_start.setCalendarPopup(True)
        self._dt_start.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        # 默认开始时间：当前时间提前5小时
        self._dt_start.setDateTime(QDateTime.currentDateTime().addSecs(-5 * 3600))
        tl.addRow("开始时间:", self._dt_start)

        self._dt_end = QDateTimeEdit()
        self._dt_end.setCalendarPopup(True)
        self._dt_end.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        # 默认结束时间：当前时间提前2.5小时（与开始时间相差2.5小时）
        self._dt_end.setDateTime(QDateTime.currentDateTime().addSecs(-2 * 3600 - 30 * 60))
        tl.addRow("结束时间:", self._dt_end)

        hb3 = QHBoxLayout()
        for label, preset in [("今天","today"),("昨天","yesterday"),
                               ("最近1h","last_1h"),("最近24h","last_24h")]:
            b = QPushButton(label)
            b.clicked.connect(lambda _, p=preset: self._set_time_range(p))
            hb3.addWidget(b)
        tl.addRow(hb3)
        
        # 自定义时间预设
        preset_layout = QHBoxLayout()
        
        self._preset_combo = QComboBox()
        self._preset_combo.setPlaceholderText("选择自定义预设...")
        self._preset_combo.currentTextChanged.connect(self._on_preset_selected)
        preset_layout.addWidget(self._preset_combo, 1)
        
        btn_manage = QPushButton("⚙️")
        btn_manage.setToolTip("管理时间预设")
        btn_manage.setFixedWidth(30)
        btn_manage.clicked.connect(self._manage_time_presets)
        preset_layout.addWidget(btn_manage)
        
        tl.addRow("自定义预设:", preset_layout)
        
        # 刷新预设列表
        self._refresh_preset_combo()

        tg.setLayout(tl)
        vbox.addWidget(tg)

        # 底部留白，与右侧版权信息对齐
        vbox.addStretch(1)

        return w

    def _make_right_panel(self) -> QWidget:
        w = QWidget()
        vbox = QVBoxLayout(w)

        # 任务表格
        task_g = QGroupBox("下载任务")
        task_l = QVBoxLayout()

        self._table = QTableWidget()
        self._table.setColumnCount(7)
        self._table.setHorizontalHeaderLabels(
            ["设备", "通道", "开始时间", "结束时间", "录像大小", "状态", "下载进度"]
        )
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_table_context_menu)
        task_l.addWidget(self._table)

        # 开始/停止下载按钮行
        hb_ctrl = QHBoxLayout()
        self._btn_start = QPushButton("▶ 开始下载")
        self._btn_start.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 6px 18px; }")
        self._btn_start.setEnabled(False)
        self._btn_start.clicked.connect(self._start_download)
        hb_ctrl.addWidget(self._btn_start)

        self._btn_stop = QPushButton("■ 停止下载")
        self._btn_stop.setStyleSheet("QPushButton { background-color: #f44336; color: white; font-weight: bold; padding: 6px 18px; }")
        self._btn_stop.setEnabled(False)
        self._btn_stop.clicked.connect(self._stop_download)
        hb_ctrl.addWidget(self._btn_stop)

        btn_clear = QPushButton("清除已完成")
        btn_clear.setStyleSheet("QPushButton { background-color: #2196F3; color: white; font-weight: bold; padding: 6px 18px; }")
        btn_clear.clicked.connect(self._clear_completed)
        hb_ctrl.addWidget(btn_clear)

        hb_ctrl.addStretch()
        task_l.addLayout(hb_ctrl)

        task_g.setLayout(task_l)
        vbox.addWidget(task_g)

        # 下载任务信息面板
        size_g = QGroupBox("📊 下载任务信息")
        size_l = QHBoxLayout()

        self._task_stats_label = QLabel("任务: 0 | 完成: 0 | 失败: 0")
        self._task_stats_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #333;")
        size_l.addWidget(self._task_stats_label)

        size_l.addWidget(self._make_separator())

        self._estimate_label = QLabel("预计大小: --")
        self._estimate_label.setStyleSheet("font-size: 13px; color: #666;")
        size_l.addWidget(self._estimate_label)

        size_l.addWidget(self._make_separator())

        self._disk_free_label = QLabel("磁盘剩余: --")
        self._disk_free_label.setStyleSheet("font-size: 13px;")
        size_l.addWidget(self._disk_free_label)

        size_l.addWidget(self._make_separator())

        self._speed_label = QLabel("下载速度: --")
        self._speed_label.setStyleSheet("font-size: 13px;")
        size_l.addWidget(self._speed_label)

        size_l.addStretch()
        size_g.setLayout(size_l)
        vbox.addWidget(size_g)

        # 日志区域（运行日志 + 下载日志 共用一个文本框）
        log_g = QGroupBox("日志")
        log_l = QVBoxLayout()
        
        # 互斥菜单切换
        chk_layout = QHBoxLayout()
        self._radio_run_log = QPushButton("📋 运行日志")
        self._radio_run_log.setCheckable(True)
        self._radio_run_log.setChecked(True)
        self._radio_run_log.setStyleSheet("QPushButton:checked { background-color: #4CAF50; color: white; }")
        self._radio_run_log.clicked.connect(lambda: self._switch_log_view("run"))
        
        self._radio_download_log = QPushButton("⬇️ 下载日志")
        self._radio_download_log.setCheckable(True)
        self._radio_download_log.setChecked(False)
        self._radio_download_log.setStyleSheet("QPushButton:checked { background-color: #2196F3; color: white; }")
        self._radio_download_log.clicked.connect(lambda: self._switch_log_view("download"))
        
        chk_layout.addWidget(self._radio_run_log)
        chk_layout.addWidget(self._radio_download_log)
        chk_layout.addStretch()
        log_l.addLayout(chk_layout)
        
        # 日志标签
        self._log_label = QLabel("运行日志:")
        log_l.addWidget(self._log_label)
        
        # 共用的日志文本框（更大的显示区域）
        self._log_text = QTextEdit()
        self._log_text.setReadOnly(True)
        self._log_text.setMinimumHeight(200)
        self._log_text.setMaximumHeight(300)
        self._log_text.setFont(QFont("Consolas", 9))
        log_l.addWidget(self._log_text)
        
        # 日志按钮行
        btn_layout = QHBoxLayout()
        btn_cl = QPushButton("清除日志")
        btn_cl.clicked.connect(self._clear_all_logs)
        btn_export = QPushButton("导出日志")
        btn_export.clicked.connect(self._export_log)
        btn_layout.addWidget(btn_cl)
        btn_layout.addWidget(btn_export)
        log_l.addLayout(btn_layout)
        
        log_g.setLayout(log_l)
        vbox.addWidget(log_g)

        return w

    # ------------------------------------------------------------------ #
    #  信号连接
    # ------------------------------------------------------------------ #

    def _wire_signals(self):
        self._progress_signal.connect(self._on_progress_ui)
        self._status_signal.connect(self._on_status_ui)
        self._log_signal.connect(self._on_log_signal)
        self._size_signal.connect(self._update_size_in_table)
        self._estimate_signal.connect(self._on_estimate_done)
        self._multi_connect_result_signal.connect(self._on_multi_connect_result)
        self._show_channel_info_signal.connect(self._on_show_channel_info)
        
        def on_progress(tid, p):
            self._progress_signal.emit(tid, p)

        def on_status(task):
            self._status_signal.emit(task.task_id)

        def on_log(msg: str):
            """日志回调（从下载器转发到GUI）"""
            self._log_signal.emit(msg)

        def on_completion(task):
            """SDK下载完成回调 → 转发到主线程日志"""
            success = task.status == DownloadStatus.COMPLETED
            QTimer.singleShot(0, lambda: self._on_task_done_bg(
                task.task_id, success, task.file_path, task.error_message
            ))

        self._dm.set_progress_callback(on_progress)
        self._dm.set_status_callback(on_status)
        self._dm.set_completion_callback(on_completion)
        self._dm.set_log_callback(on_log)
    
    def _on_log_signal(self, msg: str):
        """处理日志信号（在主线程执行）"""
        # 判断是否是下载相关日志（包含具体下载过程信息）
        download_keywords = [
            "开始下载", "下载完成", "下载失败", "下载取消",
            "开始合并", "合并完成", "合并失败", "合并取消",
            "分段", "转码", "清理", "临时", "时长", "目标", "调试日志",
            "Progress", "[Java]", "SDK", "录像", "ch", "通道", "合并模式",
            "[OK]", "[SEG]", "[SKIP]", "[WARN]", "[FAIL]", "[CONV]",
            "✓ 下载完成", "✗ 下载失败",
            # 错误相关关键词
            "错误", "error", "Error", "失败", "异常", "Exception",
            "超时", "连接", "登录", "权限", "找不到", "不存在",
            "❌", "⚠️", "[ERROR]", "[WARN]"
        ]
        
        # 运行日志关键词（这些应该显示在运行日志，而不是下载日志）
        run_keywords = [
            "▶ 开始下载", "■ 已停止", "正在连接", "连接成功", "连接失败",
            "已添加设备", "已删除设备", "保存目录", "设置已更新",
            # 设备查询相关
            "💾 获取到", "硬盘信息", "系统状态", "网络绑定", "物理网卡",
            "查询完成", "查询失败", "查询异常", "盘位", "真实IP",
            "运行时间", "工作模式", "主网卡", "从网卡", "ISAPI"
        ]
        
        # 如果匹配运行日志关键词，显示在运行日志
        is_run_log = any(kw in msg for kw in run_keywords)
        if is_run_log:
            self._log_msg(msg)
            return
        
        # 如果匹配下载关键词，显示在下载日志
        is_download_log = any(kw in msg for kw in download_keywords)
        if is_download_log:
            self._log_download(msg)
        else:
            # 其他日志显示在运行日志面板
            self._log_msg(msg)

    # ------------------------------------------------------------------ #
    #  设备管理
    # ------------------------------------------------------------------ #

    def _add_device(self):
        dlg = DeviceConfigDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            cfg = dlg.get_config()
            if not cfg['host']:
                QMessageBox.warning(self, "错误", "IP地址不能为空")
                return
            self.devices.append(cfg)
            self._save_config()
            self._refresh_device_list()
            self._log_msg(f"已添加设备: {cfg['name']} ({cfg['host']})")

    def _edit_device(self):
        row = self._device_list.currentRow()
        if row < 0:
            QMessageBox.information(self, "提示", "请先选择一个设备")
            return
        dlg = DeviceConfigDialog(self, self.devices[row])
        if dlg.exec_() == QDialog.Accepted:
            self.devices[row] = dlg.get_config()
            self._save_config()
            self._refresh_device_list()

    def _del_device(self):
        row = self._device_list.currentRow()
        if row < 0:
            return
        if QMessageBox.question(self, "确认", "确定删除该设备?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            name = self.devices.pop(row)['name']
            self._save_config()
            self._refresh_device_list()
            self._log_msg(f"已删除设备: {name}")

    def _refresh_device_list(self):
        self._device_list.clear()
        for d in self.devices:
            self._device_list.addItem(f"{d['name']}  ({d['host']}:{d.get('port',8000)})")

    def _on_device_row_changed(self, row: int):
        if row >= 0:
            self._current_config = self.devices[row]

    def _on_device_double_clicked(self, item):
        """双击设备列表项 - 连接该设备并获取通道"""
        row = self._device_list.row(item)
        if row < 0:
            return
        self._current_config = self.devices[row]
        self._device_list.setCurrentRow(row)
        self._log_msg(f"双击连接设备: {self._current_config['name']} ({self._current_config['host']})...")
        self._connect_device_by_config(self._current_config)

    def _connect_device_by_config(self, cfg: Dict):
        """根据配置连接单个设备"""
        self.statusBar().showMessage(f"连接 {cfg['host']}...")

        self._connect_worker = ConnectWorker(cfg)
        self._connect_worker.result_ready.connect(
            lambda ok, msg, dev_info, channels: self._on_connect_result_multi(ok, msg, dev_info, channels, cfg)
        )
        self._connect_worker.start()

    def _on_connect_result_multi(self, ok: bool, msg: str, dev_info: Dict, channels: List[Dict], cfg: Dict):
        """多设备连接结果处理"""
        device_key = f"{cfg['host']}:{cfg.get('port', 8000)}"

        if not ok:
            self._log_msg(f"❌ {cfg['name']} 连接失败: {msg}")
            self.statusBar().showMessage(f"{cfg['name']} 连接失败")
            return

        # 保存设备通道信息
        self._device_channels[device_key] = channels
        
        # 缓存设备信息（包括正确的通道数）
        if not hasattr(self, '_device_info_cache'):
            self._device_info_cache = {}
        self._device_info_cache[device_key] = dev_info

        # 更新通道树显示
        self._populate_channels()

        # 启用加入列表按钮（开始下载按钮在加入列表后才启用）
        self._btn_add_tasks.setEnabled(True)

        sn = dev_info.get('serial', 'Unknown')
        actual_ch = len(channels)
        online_ch = sum(1 for ch in channels if ch.get('online', True))
        offline_ch = actual_ch - online_ch

        if offline_ch > 0:
            status_str = f"通道:{actual_ch}个 (在线:{online_ch} 离线:{offline_ch})"
        else:
            status_str = f"通道:{actual_ch}个"

        self._log_msg(f"✅ {cfg['name']} 连接成功  序列号:{sn}  {status_str}")
        self.statusBar().showMessage(f"已连接 {cfg['name']} — {status_str}")

    def _query_device(self):
        """查询设备详细信息（支持多选）"""
        selected_rows = [i.row() for i in self._device_list.selectionModel().selectedRows()]

        if not selected_rows:
            QMessageBox.information(self, "提示", "请先在设备列表中选择至少一个设备（可按住Ctrl多选）")
            return

        # 支持多选查询
        for row in selected_rows:
            cfg = self.devices[row]
            self._log_msg(f"🔍 正在查询设备: {cfg['name']} ({cfg['host']})...")
            threading.Thread(target=self._do_query_device, args=(cfg,), daemon=True).start()

    def _do_query_device(self, cfg: Dict):
        """执行设备查询（在后台线程）"""
        # 辅助函数：安全地发送日志到主线程
        def log(msg: str):
            self._log_signal.emit(msg)
        
        try:
            from core.hcnetsdk import HCNetSDK
            from core.nvr_api import create_isapi

            # 使用SDK获取设备基本信息
            sdk = HCNetSDK()
            if not sdk.init():
                QTimer.singleShot(0, lambda: QMessageBox.warning(
                    self, "查询失败", f"{cfg['name']}: SDK初始化失败"
                ))
                log(f"❌ {cfg['name']}: SDK初始化失败")
                return

            # 登录获取设备信息
            ok, msg, dev = sdk.login(
                cfg['host'], cfg.get('port', 8000),
                cfg.get('username', 'admin'), cfg.get('password', '')
            )

            if not ok:
                sdk.cleanup()
                QTimer.singleShot(0, lambda: QMessageBox.warning(
                    self, "查询失败", f"{cfg['name']}: {msg}"
                ))
                log(f"❌ {cfg['name']} 查询失败: {msg}")
                return

            # 使用缓存的设备信息中的通道数（避免第二次登录计算错误）
            device_key = f"{cfg['host']}:{cfg.get('port', 8000)}"
            if hasattr(self, '_device_info_cache') and device_key in self._device_info_cache:
                cached_dev = self._device_info_cache[device_key]
                dev['total_ch'] = cached_dev['total_ch']
                log(f"  使用缓存的通道数: {dev['total_ch']}个")
            else:
                log(f"  使用登录返回的通道数: {dev['total_ch']}个")

            # 使用ISAPI获取设备信息
            hdd_info = []
            system_status = {}
            network_interfaces = []
            bond_info = {}
            try:
                api = create_isapi(cfg)
                
                # 获取硬盘信息
                hdd_info = api.get_hdd_info()
                if hdd_info:
                    log(f"  💾 获取到 {len(hdd_info)} 块硬盘信息")
                    for hdd in hdd_info:
                        hdd_id = hdd.get('id', 'N/A')
                        hdd_name = hdd.get('name', 'Unknown')
                        capacity = hdd.get('capacity', 0)
                        free = hdd.get('free', 0)
                        status = hdd.get('status', '未知')
                        serial = hdd.get('serial_number', '')
                        model = hdd.get('model', '')
                        log(f"     盘位{hdd_id}: {hdd_name} | {capacity}GB | 可用{free}GB | {status}")
                        if serial or model:
                            log(f"             序列号: {serial} | 型号: {model}")
                else:
                    log(f"  ⚠️ 未获取到硬盘信息")
                
                # 获取系统运行状态
                system_status = api.get_system_status()
                if system_status:
                    cpu = system_status.get('cpu_percent')
                    mem = system_status.get('memory_percent', 'N/A')
                    mem_usage = system_status.get('memory_usage_mb')
                    mem_total = system_status.get('memory_total_mb')
                    users = system_status.get('online_users')
                    uptime = system_status.get('uptime', 'N/A')
                    
                    # 构建CPU信息字符串（某些设备不返回CPU信息）
                    cpu_str = f"{cpu}%" if cpu is not None else "不支持"
                    
                    # 构建内存信息字符串
                    mem_str = f"{mem}%" if mem != 'N/A' else "N/A"
                    if mem_usage is not None and mem_total is not None:
                        mem_str = f"{mem}% ({mem_usage:.0f}/{mem_total:.0f}MB)"
                    
                    # 构建在线用户字符串（某些设备不返回此信息）
                    users_str = f"{users}人" if users is not None else "不支持"
                    
                    log(f"  📊 系统状态: CPU {cpu_str}, 内存 {mem_str}, 在线用户 {users_str}")
                    if uptime != 'N/A':
                        log(f"     运行时间: {uptime}")
                
                # 获取网络绑定信息（工作模式、真实IP等）
                bond_info = api.get_network_bond_info()
                if bond_info and bond_info.get('enabled'):
                    work_mode = bond_info.get('work_mode', 'N/A')
                    primary_if = bond_info.get('primary_interface', 'N/A')
                    bond_ip = bond_info.get('ip', 'N/A')
                    bond_gateway = bond_info.get('gateway', 'N/A')
                    bond_mac = bond_info.get('mac', 'N/A')
                    slaves = bond_info.get('slave_interfaces', [])
                    
                    log(f"  🔗 网络绑定: 工作模式={work_mode}")
                    log(f"     真实IP: {bond_ip} | 网关: {bond_gateway}")
                    log(f"     主网卡: Lan{primary_if} | 从网卡: {[f'Lan{s}' for s in slaves]}")
                
                # 获取网络接口信息（物理网卡）
                network_interfaces = api.get_network_interfaces()
                if network_interfaces:
                    log(f"  🌐 物理网卡: {len(network_interfaces)} 个")
                    for iface in network_interfaces:
                        iface_id = iface.get('id', 'N/A')
                        ip = iface.get('ip', 'N/A')
                        mac = iface.get('mac', 'N/A')
                        log(f"     Lan{iface_id}: {ip} | MAC:{mac}")
                
            except Exception as e:
                log(f"  ❌ 获取设备信息失败: {e}")

            sdk.cleanup()

            # 构建查询结果
            result_text = f"""📋 设备查询结果: {cfg['name']}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📍 基本信息
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  设备名称: {cfg['name']}
  IP地址:   {cfg['host']}:{cfg.get('port', 8000)}
  序列号:   {dev.get('serial', 'Unknown')}
  设备类型: {dev.get('type', 'Unknown')}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📊 通道信息
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  模拟通道: {dev.get('analog_ch', 0)} 个
  IP通道:   {dev.get('ip_ch', 0)} 个
  总通道:   {dev.get('total_ch', 0)} 个
  起始通道: {dev.get('start_chan', 1)}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
💾 硬盘信息
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""

            if hdd_info:
                for hdd in hdd_info:
                    hdd_id = hdd.get('id', 'N/A')
                    hdd_name = hdd.get('name', '')
                    status = hdd.get('status', '未知')
                    capacity = hdd.get('capacity', 0)
                    free = hdd.get('free', 0)
                    used = capacity - free
                    serial = hdd.get('serial_number', '')
                    model = hdd.get('model', '')
                    name_str = f" ({hdd_name})" if hdd_name else ""
                    result_text += f"\n  盘位{hdd_id}{name_str}: {status} | 总容量: {capacity}GB | 已用: {used}GB | 可用: {free}GB"
                    if serial or model:
                        result_text += f"\n           序列号: {serial} | 型号: {model}"
            else:
                result_text += "\n  (未获取到硬盘信息)"

            result_text += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"

            # 运行状态信息
            result_text += """

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📈 运行状态
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            if system_status:
                cpu = system_status.get('cpu_percent')
                mem = system_status.get('memory_percent', 'N/A')
                mem_usage = system_status.get('memory_usage_mb')
                mem_total = system_status.get('memory_total_mb')
                users = system_status.get('online_users')
                uptime = system_status.get('uptime', 'N/A')
                
                # 构建内存信息字符串
                mem_str = f"{mem}%"
                if mem_usage is not None and mem_total is not None:
                    mem_str = f"{mem}% ({mem_usage:.0f}/{mem_total:.0f}MB)"
                
                # CPU使用率（某些设备不返回）
                if cpu is not None:
                    result_text += f"\n  CPU使用率: {cpu}%"
                
                result_text += f"\n  内存使用率: {mem_str}"
                
                # 在线用户（某些设备不返回）
                if users is not None:
                    result_text += f"\n  在线用户: {users} 人"
                
                if uptime != 'N/A':
                    result_text += f"\n  运行时间: {uptime}"
            else:
                result_text += "\n  (未获取到运行状态)"

            # 网络绑定信息（Bond）
            result_text += """

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🔗 网络配置
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            if bond_info and bond_info.get('enabled'):
                work_mode = bond_info.get('work_mode', 'N/A')
                primary_if = bond_info.get('primary_interface', 'N/A')
                bond_ip = bond_info.get('ip', 'N/A')
                bond_mask = bond_info.get('mask', 'N/A')
                bond_gateway = bond_info.get('gateway', 'N/A')
                bond_mac = bond_info.get('mac', 'N/A')
                slaves = bond_info.get('slave_interfaces', [])
                
                result_text += f"\n  工作模式: {work_mode}"
                result_text += f"\n  真实IP地址: {bond_ip}"
                result_text += f"\n  子网掩码: {bond_mask}"
                if bond_gateway != 'N/A':
                    result_text += f"\n  网关: {bond_gateway}"
                if bond_mac != 'N/A':
                    result_text += f"\n  MAC地址: {bond_mac}"
                result_text += f"\n  主网卡: Lan{primary_if}"
                if slaves:
                    result_text += f"\n  从网卡: {', '.join([f'Lan{s}' for s in slaves])}"
            else:
                result_text += "\n  (未启用网络绑定)"

            # 物理网卡信息
            result_text += """

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🌐 物理网卡
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
            if network_interfaces:
                for iface in network_interfaces:
                    iface_id = iface.get('id', 'N/A')
                    ip = iface.get('ip', 'N/A')
                    mask = iface.get('mask', 'N/A')
                    mac = iface.get('mac', 'N/A')
                    mtu = iface.get('mtu', 'N/A')
                    
                    result_text += f"\n  Lan{iface_id}:"
                    result_text += f"\n    IP地址: {ip}"
                    result_text += f"\n    子网掩码: {mask}"
                    if mac != 'N/A':
                        result_text += f"\n    MAC地址: {mac}"
                    if mtu != 'N/A':
                        result_text += f"\n    MTU: {mtu}"
            else:
                result_text += "\n  (未获取到物理网卡信息)"

            result_text += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"

            # 显示结果对话框
            QTimer.singleShot(0, lambda: QMessageBox.information(
                self, f"设备查询 - {cfg['name']}", result_text
            ))

            # 日志中同时显示硬盘信息和序列号
            hdd_count = len(hdd_info) if hdd_info else 0
            log(f"✅ {cfg['name']} 查询完成: 序列号:{dev.get('serial', 'Unknown')}, 通道:{dev.get('total_ch', 0)}个, 硬盘:{hdd_count}块")

        except Exception as e:
            QTimer.singleShot(0, lambda: QMessageBox.warning(
                self, "查询异常", f"{cfg['name']}: {str(e)}"
            ))
            log(f"❌ {cfg['name']} 查询异常: {e}")

    def _on_query_channel_info_clicked(self):
        """菜单点击回调"""
        self._query_channel_info()
        
    def _query_channel_info(self):
        """查询通道详细信息（分辨率、码率、编码等）"""
        selected_rows = [i.row() for i in self._device_list.selectionModel().selectedRows()]
        
        if not selected_rows:
            QMessageBox.information(self, "提示", "请先在设备列表中选择一台设备")
            return
        
        if len(selected_rows) > 1:
            QMessageBox.information(self, "提示", "请只选择一台设备进行通道信息查询")
            return
        
        cfg = self.devices[selected_rows[0]]
        self._log_msg(f"🔍 正在查询通道信息: {cfg['name']} ({cfg['host']})...")
        
        def wrapped_query(cfg):
            try:
                self._do_query_channel_info(cfg)
            except Exception as e:
                import traceback
                self._log_msg(f"查询通道信息失败: {e}")
                self._log_msg(f"详情: {traceback.format_exc()}")
        
        t = threading.Thread(target=wrapped_query, args=(cfg,), daemon=True)
        t.start()

    def _do_query_channel_info(self, cfg: Dict):
        """执行通道信息查询（在后台线程）"""
        # 使用信号发送日志到主线程
        def log(msg: str):
            self._log_signal.emit(msg)
        
        try:
            from core.nvr_api import create_isapi
            
            api = create_isapi({
                'host': cfg['host'],
                'http_port': cfg.get('http_port', 80),
                'username': cfg.get('username', 'admin'),
                'password': cfg.get('password', ''),
            })
            
            # 获取通道流信息
            stream_info = api.get_channel_stream_info()
            log(f"  [DEBUG] 获取到 {len(stream_info)} 个通道的流信息: {list(stream_info.keys())}")
            
            if not stream_info:
                log(f"⚠️ {cfg['name']}: 未获取到通道流信息")
                QTimer.singleShot(0, lambda: QMessageBox.warning(
                    self, "查询失败", f"{cfg['name']}: 未获取到通道流信息"
                ))
                return
            
            # 获取通道名称和在线状态
            channels_with_status = api.get_channels_with_status()
            
            # 准备表格数据
            table_data = []
            
            for ch_no in sorted(stream_info.keys()):
                try:
                    info = stream_info[ch_no]
                    main_stream = info.get('main_stream', {})
                    sub_stream = info.get('sub_stream', {})
                    
                    # DEBUG: 打印前几个通道的流信息详情
                    if ch_no <= 5:
                        log(f"  [DEBUG] 通道{ch_no}: main_stream={main_stream}, sub_stream={sub_stream}")
                    ch_status = channels_with_status.get(ch_no, {})
                    
                    # 构建编码格式显示
                    def build_codec_display(stream):
                        if not stream:
                            return "-"  # 无流信息（通道离线或未配置码流）
                        codec = stream.get('codec', 'N/A')
                        if codec == 'N/A':
                            return "N/A"  # 有流信息但无法识别编码
                        codec_profile = stream.get('codec_profile', '')
                        smart_codec = stream.get('smart_codec', False)
                        smart_codec_type = stream.get('smart_codec_type', '')
                        
                        display = codec
                        if codec_profile:
                            display += f" ({codec_profile})"
                        if smart_codec:
                            smart_type_str = f" - {smart_codec_type}" if smart_codec_type else ""
                            display += f" [Smart{smart_type_str}]"
                        return display
                    
                    # 获取OSD名称 - 对于此NVR，OSD名称就是通道名称
                    osd_name = ch_status.get('name', f'通道{ch_no}')
                    
                    row = {
                        'channel_no': ch_no,
                        'channel_name': ch_status.get('name', f'通道{ch_no}'),
                        'online': '在线' if ch_status.get('online', False) else '离线',
                        'ip': ch_status.get('ip', ''),
                        'protocol': ch_status.get('protocol', '').upper(),
                        'osd_name': osd_name,
                        'main_resolution': main_stream.get('resolution', 'N/A') if main_stream else '未配置',
                        'main_codec': build_codec_display(main_stream),
                        'main_bitrate_mode': main_stream.get('bitrate_mode', '-') if main_stream else '-',
                        'main_bitrate': str(main_stream.get('bitrate_kbps', 0)) + ' kbps' if main_stream and main_stream.get('bitrate_kbps', 0) > 0 else '-',
                        'main_fps': str(main_stream.get('fps', 0)) if main_stream and main_stream.get('fps', 0) > 0 else '-',
                        'sub_resolution': sub_stream.get('resolution', 'N/A') if sub_stream else '未配置',
                        'sub_codec': build_codec_display(sub_stream),
                        'sub_bitrate_mode': sub_stream.get('bitrate_mode', '-') if sub_stream else '-',
                        'sub_bitrate': str(sub_stream.get('bitrate_kbps', 0)) + ' kbps' if sub_stream and sub_stream.get('bitrate_kbps', 0) > 0 else '-',
                        'sub_fps': str(sub_stream.get('fps', 0)) if sub_stream and sub_stream.get('fps', 0) > 0 else '-',
                    }
                    table_data.append(row)
                except Exception as row_e:
                    log(f"⚠️ 处理通道 {ch_no} 时出错: {row_e}")
            
            if table_data:
                self._show_channel_info_signal.emit(cfg['name'], table_data)
            else:
                log(f"⚠️ {cfg['name']}: 表格数据为空，无法显示对话框")
            
            log(f"✅ {cfg['name']}: 获取到 {len(stream_info)} 个通道的流信息")
            
        except Exception as e:
            import traceback
            log(f"❌ {cfg['name']} 查询通道信息异常: {e}")
            QTimer.singleShot(0, lambda: QMessageBox.warning(
                self, "查询异常", f"{cfg['name']}: {str(e)}"
            ))

    def _on_show_channel_info(self, device_name: str, table_data: list):
        """信号处理：显示通道信息表格对话框"""
        try:
            # 查找设备配置
            device_config = next((dev for dev in self.devices if dev['name'] == device_name), None)
            
            dialog = ChannelInfoDialog(device_name, table_data, device_config, self)
            # 连接OSD更新信号
            dialog.osd_update_signal.connect(self._on_osd_update_requested)
            dialog.exec_()
        except Exception as e:
            import traceback
            self._log_msg(f"❌ 显示通道信息对话框失败: {e}")
            QMessageBox.warning(self, "显示失败", f"无法显示通道信息对话框:\n{str(e)}")
    
    def _on_osd_update_requested(self, osd_updates: list, device_config: dict):
        """处理OSD更新请求"""
        self._log_msg(f"📝 收到OSD更新请求: {device_config.get('name', 'Unknown')}，共 {len(osd_updates)} 个通道")
        
        # 启动后台线程进行批量更新
        threading.Thread(
            target=self._do_batch_osd_update,
            args=(osd_updates, device_config),
            daemon=True
        ).start()
    
    def _do_batch_osd_update(self, osd_updates: list, device_config: dict):
        """执行批量OSD更新（后台线程）"""
        def log(msg: str):
            self._log_signal.emit(msg)
        
        try:
            from core.nvr_api import create_isapi
            
            log(f"[OSD] 开始批量更新，设备: {device_config.get('name', 'Unknown')}")
            
            api = create_isapi({
                'host': device_config['host'],
                'http_port': device_config.get('http_port', 80),
                'username': device_config.get('username', 'admin'),
                'password': device_config.get('password', ''),
            })
            
            success_count = 0
            fail_count = 0
            
            for i, (ch_no, osd_name) in enumerate(osd_updates, 1):
                log(f"[OSD] [{i}/{len(osd_updates)}] 更新通道 {ch_no}: '{osd_name}'")
                
                try:
                    success, msg = api.set_channel_osd(ch_no, osd_name)
                    if success:
                        success_count += 1
                        log(f"✅ 通道{ch_no}: {msg}")
                    else:
                        fail_count += 1
                        log(f"❌ 通道{ch_no}: {msg}")
                except Exception as e:
                    fail_count += 1
                    log(f"❌ 通道{ch_no}: 调用异常 - {str(e)}")
                
                # 添加小延迟避免请求过快
                import time
                time.sleep(0.2)
            
            log(f"[OSD] 批量更新完成: ✅成功 {success_count} 个, ❌失败 {fail_count} 个")
            
            # 显示完成提示
            QTimer.singleShot(0, lambda: QMessageBox.information(
                self, "OSD更新完成",
                f"设备: {device_config.get('name', 'Unknown')}\n\n"
                f"✅ 成功: {success_count} 个\n"
                f"❌ 失败: {fail_count} 个\n\n"
                f"请重新查询通道信息以查看更新结果。"
            ))
            
        except Exception as e:
            import traceback
            log(f"❌ [OSD] 批量更新异常: {e}")
            QTimer.singleShot(0, lambda: QMessageBox.warning(
                self, "OSD更新失败", f"批量更新OSD时发生错误:\n{str(e)}"
            ))

    # ------------------------------------------------------------------ #
    #  连接 / 刷新通道
    # ------------------------------------------------------------------ #

    def _connect_device(self):
        row = self._device_list.currentRow()
        if row < 0:
            QMessageBox.information(self, "提示", "请先在列表中选择设备")
            return
        self._current_config = self.devices[row]
        self._log_msg(f"正在连接 {self._current_config['host']}...")
        self.statusBar().showMessage("连接中...")
        self._channel_tree.clear()
        self._channel_count_label.setText("获取中...")

        # 启动后台连接线程
        self._connect_worker = ConnectWorker(self._current_config)
        self._connect_worker.result_ready.connect(self._on_connect_result)
        self._connect_worker.start()
        self._log_msg(f"✓ 连接线程已启动，等待响应...")

    def _refresh_channels(self):
        """刷新通道 - 支持多选设备同时刷新"""
        selected_rows = [i.row() for i in self._device_list.selectionModel().selectedRows()]

        if not selected_rows:
            QMessageBox.information(self, "提示", "请先在设备列表中选择至少一个设备（可按住Ctrl多选）")
            return

        if len(selected_rows) == 1:
            # 单选，使用原有逻辑
            self._current_config = self.devices[selected_rows[0]]
            self._connect_device()
        else:
            # 多选，使用线程池同时连接多个设备
            self._log_msg(f"▶ 开始批量连接 {len(selected_rows)} 台设备...")
            self._connect_multiple_devices([self.devices[row] for row in selected_rows])

    def _connect_multiple_devices(self, configs: List[Dict]):
        """使用线程池同时连接多个设备"""
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading

        def connect_single(cfg: Dict) -> tuple:
            """连接单个设备并返回结果"""
            try:
                from core.hcnetsdk import HCNetSDK
                sdk = HCNetSDK()
                if not sdk.init():
                    return cfg, False, "SDK初始化失败", {}, []

                ok, msg, dev = sdk.login(
                    cfg['host'], cfg.get('port', 8000),
                    cfg.get('username', 'admin'), cfg.get('password', '')
                )
                if not ok:
                    sdk.cleanup()
                    return cfg, False, f"登录失败: {msg}", {}, []

                channels = sdk.get_channels_with_names(
                    total_ch=dev['total_ch'],
                    start_chan=max(dev.get('start_dchan', 1), 1),
                    nvr_ip=cfg['host'],
                    nvr_port=cfg.get('http_port', 80),
                    username=cfg.get('username', 'admin'),
                    password=cfg.get('password', ''),
                )
                sdk.cleanup()
                return cfg, True, "连接成功", dev, channels
            except Exception as e:
                return cfg, False, f"异常: {e}", {}, []

        def run_connections():
            with ThreadPoolExecutor(max_workers=min(len(configs), 5)) as executor:
                futures = {executor.submit(connect_single, cfg): cfg for cfg in configs}
                for future in as_completed(futures):
                    cfg, ok, msg, dev, channels = future.result()
                    # 使用信号发送到主线程更新UI
                    self._multi_connect_result_signal.emit(cfg, ok, msg, dev, channels)

        threading.Thread(target=run_connections, daemon=True).start()

    def _on_multi_connect_result(self, cfg: Dict, ok: bool, msg: str, dev: Dict, channels: List[Dict]):
        """批量连接结果处理（在主线程执行）"""
        device_key = f"{cfg['host']}:{cfg.get('port', 8000)}"

        if not ok:
            self._log_msg(f"❌ {cfg['name']} 连接失败: {msg}")
        else:
            self._device_channels[device_key] = channels
            self._populate_channels()

            # 启用加入列表按钮（开始下载按钮在加入列表后才启用）
            self._btn_add_tasks.setEnabled(True)

            sn = dev.get('serial', 'Unknown')
            actual_ch = len(channels)
            online_ch = sum(1 for ch in channels if ch.get('online', True))
            offline_ch = actual_ch - online_ch

            if offline_ch > 0:
                status_str = f"通道:{actual_ch}个 (在线:{online_ch} 离线:{offline_ch})"
            else:
                status_str = f"通道:{actual_ch}个"

            self._log_msg(f"✅ {cfg['name']} 连接成功  序列号:{sn}  {status_str}")
            self.statusBar().showMessage(f"已连接 {cfg['name']} — {status_str}")

    def _on_connect_result(self, ok: bool, msg: str, dev_info: Dict, channels: List[Dict]):
        if not ok:
            self._log_msg(f"❌ 连接失败: {msg}")
            self.statusBar().showMessage(f"连接失败: {msg}")
            QMessageBox.warning(self, "连接失败", msg)
            return

        # 保存设备通道信息
        device_key = f"{self._current_config['host']}:{self._current_config.get('port', 8000)}"
        self._device_channels[device_key] = channels

        # 更新通道树显示（支持多设备）
        self._populate_channels()
        # 启用加入列表按钮（开始下载按钮在加入列表后才启用）
        self._btn_add_tasks.setEnabled(True)

        sn = dev_info.get('serial', '')
        # 优先显示ISAPI实际通道数，而非SDK返回的通道数（SDK可能返回设备最大容量128而非实际接入数）
        actual_ch = len(channels)
        online_ch = sum(1 for ch in channels if ch.get('online', True))
        offline_ch = actual_ch - online_ch

        if offline_ch > 0:
            status_str = f"通道:{actual_ch}个 (在线:{online_ch} 离线:{offline_ch})"
        else:
            status_str = f"通道:{actual_ch}个"

        self._log_msg(f"✅ 连接成功  序列号:{sn}  {status_str}")
        self.statusBar().showMessage(f"已连接 {self._current_config['host']} — {status_str}")

    def _populate_channels(self):
        """填充通道树（支持多设备分组，离线通道标注）"""
        self._channel_tree.clear()

        total_channels = 0
        for device in self.devices:
            device_key = f"{device['host']}:{device.get('port', 8000)}"
            channels = self._device_channels.get(device_key, [])

            if not channels:
                continue

            # 统计在线/离线数
            online_count  = sum(1 for ch in channels if ch.get('online', True))
            offline_count = len(channels) - online_count

            # 创建设备节点
            device_item = QTreeWidgetItem(self._channel_tree)
            device_name = device.get('name', device['host'])
            if offline_count > 0:
                device_item.setText(0, f"📹 {device_name} ({len(channels)}通道, {offline_count}个离线)")
            else:
                device_item.setText(0, f"📹 {device_name} ({len(channels)}通道, 全部在线)")
            device_item.setExpanded(True)  # 默认展开
            device_item.setCheckState(0, Qt.Unchecked)  # 设备项也带checkbox，用于批量选择

            # 添加通道子节点
            for ch in channels:
                ch_item = QTreeWidgetItem(device_item)
                ch_name = ch.get('name', f"通道{ch['id']}")
                is_online = ch.get('online', True)

                if is_online:
                    # 在线通道：正常显示，名称前加上序号
                    ch_item.setText(0, f"{ch['id']}. {ch_name}")
                    ch_item.setCheckState(0, Qt.Unchecked)
                    ch_item.setForeground(0, ch_item.foreground(0))  # 默认颜色
                else:
                    # 离线通道：灰色显示，名称前加上序号，名称后加 [离线] 标注
                    ch_item.setText(0, f"{ch['id']}. {ch_name}  [离线]")
                    ch_item.setCheckState(0, Qt.Unchecked)
                    from PyQt5.QtGui import QColor
                    ch_item.setForeground(0, QColor(150, 150, 150))  # 灰色
                    ch_item.setToolTip(0, f"通道离线 (status: {ch.get('status', 'unknown')})")

                ch_item.setData(0, Qt.UserRole, {**ch, 'device': device})

            total_channels += len(channels)

        self._update_main_channel_count()

    def _update_main_channel_count(self):
        """更新主窗口通道数量显示（总数和已选择数）"""
        total = 0
        selected = 0
        for i in range(self._channel_tree.topLevelItemCount()):
            device_item = self._channel_tree.topLevelItem(i)
            for j in range(device_item.childCount()):
                total += 1
                if device_item.child(j).checkState(0) == Qt.Checked:
                    selected += 1
        self._channel_count_label.setText(f"共 {total} 个通道，已选择 {selected} 个")

    def _select_all(self):
        """全选通道：优先全选被选中的设备下的通道，若无选中设备则全选所有"""
        selected_devices = []
        for i in range(self._channel_tree.topLevelItemCount()):
            device_item = self._channel_tree.topLevelItem(i)
            if device_item.checkState(0) == Qt.Checked:
                selected_devices.append(device_item)

        # 如果有选中的设备，只全选这些设备下的通道
        target_devices = selected_devices if selected_devices else [
            self._channel_tree.topLevelItem(i) for i in range(self._channel_tree.topLevelItemCount())
        ]

        for device_item in target_devices:
            for j in range(device_item.childCount()):
                device_item.child(j).setCheckState(0, Qt.Checked)
        self._update_main_channel_count()

    def _deselect_all(self):
        """取消全选：优先取消被选中的设备下的通道，若无选中设备则取消所有"""
        selected_devices = []
        for i in range(self._channel_tree.topLevelItemCount()):
            device_item = self._channel_tree.topLevelItem(i)
            if device_item.checkState(0) == Qt.Checked:
                selected_devices.append(device_item)

        # 如果有选中的设备，只取消这些设备下的通道
        target_devices = selected_devices if selected_devices else [
            self._channel_tree.topLevelItem(i) for i in range(self._channel_tree.topLevelItemCount())
        ]

        for device_item in target_devices:
            for j in range(device_item.childCount()):
                device_item.child(j).setCheckState(0, Qt.Unchecked)
        self._update_main_channel_count()

    def _get_selected_channels(self) -> List[Dict]:
        """获取所有选中的通道（支持多设备）"""
        result = []
        for i in range(self._channel_tree.topLevelItemCount()):
            device_item = self._channel_tree.topLevelItem(i)
            for j in range(device_item.childCount()):
                ch_item = device_item.child(j)
                if ch_item.checkState(0) == Qt.Checked:
                    ch = ch_item.data(0, Qt.UserRole)
                    if ch:
                        result.append(ch)
        return result

    # ------------------------------------------------------------------ #
    #  流信息预加载缓存
    # ------------------------------------------------------------------ #

    def _add_tasks_to_list(self):
        """将选中通道加入下载列表"""
        selected = self._get_selected_channels()
        if not selected:
            QMessageBox.warning(self, "提示", "请先勾选至少一个通道")
            return

        start_dt = self._dt_start.dateTime().toPyDateTime()
        end_dt   = self._dt_end.dateTime().toPyDateTime()
        if start_dt >= end_dt:
            QMessageBox.warning(self, "提示", "开始时间必须早于结束时间")
            return

        import uuid
        added_count = 0
        device_groups = {}

        # ISAPI模式：先同步探测每个任务的大小，再加入列表
        # 这样加入时就能显示实际大小
        device_probe_results: Dict[str, int] = {}  # {task_id: size_bytes}

        if self._download_mode == "isapi":
            self._estimate_label.setText("预计大小: 探测中...")
            self._estimate_label.setStyleSheet("font-size: 13px; color: #FF9800;")
        # elif self._download_mode == "hikload":  # HikLoad模式已禁用
        #     self._estimate_label.setText("预计大小: RTSP流式下载")
        #     self._estimate_label.setStyleSheet("font-size: 13px; color: #2196F3;")

        # 第一步：收集所有任务信息（先创建任务对象用于关联task_id）
        new_tasks: List[Tuple[DownloadTask, Dict, Dict]] = []  # (task, device_config, channel)
        for ch in selected:
            device = ch.get('device') or self._current_config
            if not device:
                continue
            device_key = f"{device['host']}:{device.get('port', 8000)}"

            if device_key not in self._device_channels:
                QMessageBox.warning(self, "提示",
                    f"设备 {device.get('name', device['host'])} 未连接，请先连接")
                return

            if device_key not in device_groups:
                device_groups[device_key] = {
                    'config': device,
                    'channels': []
                }
            device_groups[device_key]['channels'].append(ch)

            # 创建临时任务对象（仅用于生成task_id）
            task = DownloadTask(
                task_id      = str(uuid.uuid4()),
                device_id    = device_key,
                device_name  = device_groups[device_key]['config'].get('name', device_groups[device_key]['config']['host']),
                device_config = device_groups[device_key]['config'],
                channel_id   = str(ch.get('no', ch.get('id', '1'))),
                channel_name = ch.get('name', f"通道{ch.get('id','?')}"),
                start_time   = start_dt,
                end_time     = end_dt,
                save_dir     = self.download_dir,
                merge_mode   = getattr(self, '_merge_mode', 'standard'),
                enable_debug_log = getattr(self, '_enable_debug_log', False),
                skip_transcode = getattr(self, '_skip_transcode', True),
            )
            new_tasks.append((task, device_groups[device_key]['config'], ch))

        # 第二步：ISAPI同步探测每个任务大小（最多重试5次）
        # - 离线设备跳过探测，表格显示"离线"
        # - 在线设备探测失败则重试，最多重试5次
        MAX_RETRIES = 5
        offline_task_ids = []  # 离线任务列表

        if self._download_mode == "isapi" and new_tasks:
            from core.nvr_api import create_isapi
            from PyQt5.QtCore import QMetaObject, Qt
            
            # 调试：打印设备配置
            if new_tasks:
                _, first_dev_cfg, _ = new_tasks[0]
                print(f"[PROBE DEBUG] 设备配置: {first_dev_cfg}")

            for task, dev_cfg, ch in new_tasks:
                # 跳过离线通道（online 是布尔值 True/False）
                if not ch.get('online', True):
                    offline_task_ids.append(task.task_id)
                    continue

                size = 0
                for attempt in range(1, MAX_RETRIES + 1):
                    try:
                        api = create_isapi(dev_cfg)
                        ch_no = int(ch.get('no', ch.get('id', 1)))
                        size = api.probe_record_size(
                            channel=ch_no,
                            start_time=start_dt,
                            end_time=end_dt,
                            stream_type=1,
                            rtsp_port=dev_cfg.get('rtsp_port', 554),
                        )
                        if size > 0:
                            break  # 成功，退出重试循环
                        elif attempt == 1:
                            print(f"[PROBE] 通道{ch_no} 探测返回 size={size}（可能无录像或接口不支持）")
                    except Exception as e:
                        if attempt == 1:
                            print(f"[PROBE] 通道{ch_no} 探测异常: {e}")
                        import traceback
                        traceback.print_exc()

                    if attempt < MAX_RETRIES:
                        time.sleep(0.3)  # 重试前等待300ms

                if size > 0:
                    device_probe_results[task.task_id] = size
                    self._task_file_sizes[task.task_id] = size
                else:
                    # 5次探测全部失败，标记为探测失败
                    device_probe_results[task.task_id] = -1  # -1 表示探测失败

        # 第三步：正式加入列表
        # offline_task_ids 中的任务显示"离线"（橙色）
        # device_probe_results 中值为 -1 的任务显示"探测失败"（红色）
        # 值为 > 0 的任务显示实际大小
        # 其余（=0）显示"探测中..."（灰色）
        for task, _, _ in new_tasks:
            self._pending_tasks.append(task)
            if task.task_id in offline_task_ids:
                self._add_row(task, -2)  # -2 = 离线
            else:
                self._add_row(task, device_probe_results.get(task.task_id, 0))
            added_count += 1

        # 启用开始下载按钮
        if self._pending_tasks:
            self._btn_start.setEnabled(True)

        # 更新磁盘信息
        self._update_disk_info()

        device_count = len(device_groups)
        probed = sum(1 for v in device_probe_results.values() if v > 0)
        failed = sum(1 for v in device_probe_results.values() if v == -1)
        offline = len(offline_task_ids)

        # 更新预计大小标签（与日志一致）
        if self._download_mode == "isapi":
            if probed > 0:
                total_probed_size = sum(v for v in device_probe_results.values() if v > 0)
                parts = [f"预计大小: {self._format_bytes(total_probed_size)}"]
                if failed > 0:
                    parts.append(f"（{failed}个探测失败）")
                if offline > 0:
                    parts.append(f"（{offline}个离线）")
                self._estimate_label.setText(" ".join(parts))
                self._estimate_label.setStyleSheet("font-size: 13px; color: #333; font-weight: bold;")
            else:
                self._estimate_label.setText("预计大小: 探测失败")
                self._estimate_label.setStyleSheet("font-size: 13px; color: #f44336;")
        # elif self._download_mode == "hikload":  # HikLoad模式已禁用
        #     self._estimate_label.setText("预计大小: RTSP流式下载")
        #     self._estimate_label.setStyleSheet("font-size: 13px; color: #2196F3;")
        else:
            self._estimate_label.setText("预计大小: --")
            self._estimate_label.setStyleSheet("font-size: 13px; color: #666;")

        # 日志消息
        if self._download_mode == "isapi":
            if probed > 0:
                parts = [f"已探测 {probed}/{len(new_tasks)} 个，大小合计 {self._format_bytes(total_probed_size)}"]
                if failed > 0:
                    parts.append(f"（{failed}个探测失败，设备可能不支持ISAPI下载）")
                if offline > 0:
                    parts.append(f"（{offline}个离线）")
                self._log_msg(f"📋 已加入 {added_count} 个任务（{device_count}台设备），" + "，".join(parts))
            else:
                # 全部探测失败，提示用户切换模式
                msg = f"📋 已加入 {added_count} 个任务（{device_count}台设备）"
                if failed > 0:
                    msg += f"，{failed}个探测失败"
                    # 检查是否所有任务都探测失败
                    if failed == len(new_tasks) - offline:
                        msg += "（设备可能不支持ISAPI下载，请尝试切换到SDK模式）"
                if offline > 0:
                    msg += f"，{offline}个离线"
                self._log_msg(msg + "，点击「开始下载」执行")
        # elif self._download_mode == "hikload":  # HikLoad模式已禁用
        #     self._log_msg(f"📋 已加入 {added_count} 个任务（{device_count}台设备）- RTSP流式下载模式，点击「开始下载」执行")
        else:
            self._log_msg(f"📋 已加入 {added_count} 个任务（{device_count}台设备），点击「开始下载」执行")

    # ------------------------------------------------------------------ #
    #  录像大小探测
    # ------------------------------------------------------------------ #

    def _probe_sizes_worker(self, probe_tasks, start_dt, end_dt):
        """后台线程：逐个探测通道录像大小，汇总后通过信号通知主线程"""
        from core.nvr_api import create_isapi

        total_size = 0
        ok_count = 0
        fail_count = 0

        for task_id, device_config, ch in probe_tasks:
            try:
                api = create_isapi(device_config)
                ch_no = int(ch.get('no', ch.get('id', 1)))
                size = api.probe_record_size(
                    channel=ch_no,
                    start_time=start_dt,
                    end_time=end_dt,
                    stream_type=1,
                    rtsp_port=device_config.get('rtsp_port', 554),
                )
                if size > 0:
                    total_size += size
                    ok_count += 1
                    # 同时更新表格中该任务的大小列
                    self._task_file_sizes[task_id] = size
                    self._size_signal.emit(task_id, size)
                else:
                    fail_count += 1
            except Exception:
                fail_count += 1

        self._estimate_signal.emit(total_size, ok_count, fail_count)

    def _on_estimate_done(self, total_size: int, ok_count: int, fail_count: int):
        """主线程槽：探测完成后更新预计大小标签"""
        if total_size > 0:
            size_str = self._format_bytes(total_size)
            parts = [f"预计大小: {size_str}"]
            if fail_count > 0:
                parts.append(f"（{ok_count}成功 {fail_count}失败）")
            self._estimate_label.setText(" ".join(parts))
            self._estimate_label.setStyleSheet("font-size: 13px; color: #333; font-weight: bold;")
        else:
            self._estimate_label.setText("预计大小: 探测失败")
            self._estimate_label.setStyleSheet("font-size: 13px; color: #f44336;")

    def _make_separator(self) -> QWidget:
        """创建竖线分隔符"""
        sep = QWidget()
        sep.setFixedWidth(1)
        sep.setStyleSheet("background-color: #ccc;")
        return sep


    # ------------------------------------------------------------------ #
    #  磁盘信息 + 下载速度
    # ------------------------------------------------------------------ #

    def _update_disk_info(self):
        """更新磁盘剩余空间显示"""
        try:
            import ctypes
            free_bytes = ctypes.c_ulonglong(0)
            total_bytes = ctypes.c_ulonglong(0)
            drive = os.path.splitdrive(self.download_dir)[0] + "\\"
            ctypes.windll.kernel32.GetDiskFreeSpaceExW(
                drive, None, ctypes.pointer(total_bytes), ctypes.pointer(free_bytes)
            )
            free_gb = free_bytes.value / (1024**3)
            total_gb = total_bytes.value / (1024**3)
            self._disk_free_label.setStyleSheet("font-size: 13px; color: #333;")
            self._disk_free_label.setText(f"磁盘剩余: {free_gb:.1f} GB / {total_gb:.1f} GB")
        except Exception as e:
            self._disk_free_label.setText(f"磁盘剩余: 获取失败")
            self._disk_free_label.setStyleSheet("font-size: 13px; color: #f44336;")







    def _update_download_speed(self):
        """使用网卡速度显示下载速度"""
        now = time.time()
        total_speed = 0
        active_count = 0

        try:
            import psutil
            # 获取网卡速度（发送+接收）
            net_io = psutil.net_io_counters()
            if self._last_net_io is not None and self._last_net_time > 0:
                elapsed = now - self._last_net_time
                if elapsed > 0:
                    bytes_sent = net_io.bytes_sent - self._last_net_io.bytes_sent
                    bytes_recv = net_io.bytes_recv - self._last_net_io.bytes_recv
                    total_bytes = bytes_sent + bytes_recv
                    total_speed = total_bytes / elapsed
                    # 检查是否有活跃下载任务（下载中或合并中）
                    for task_id in list(self._download_start_times.keys()):
                        task = self._dm.get_task(task_id)
                        if task and task.status in (DownloadStatus.DOWNLOADING, DownloadStatus.MERGING):
                            active_count += 1
            self._last_net_io = net_io
            self._last_net_time = now
        except Exception as e:
            print(f"[速度监控] 异常: {e}")

        if active_count > 0 and total_speed > 0:
            speed_str = self._format_speed(total_speed)
            self._speed_label.setText(f"网卡速度: {speed_str} ({active_count}任务)")
            self._speed_label.setStyleSheet("font-size: 13px; color: #2196F3; font-weight: bold;")
        elif active_count > 0:
            self._speed_label.setText(f"网卡速度: 计算中... ({active_count}任务)")
            self._speed_label.setStyleSheet("font-size: 13px; color: #999;")
        else:
            self._speed_label.setText("网卡速度: --")
            self._speed_label.setStyleSheet("font-size: 13px; color: #999;")

        # 周期性检查：为已完成但大小仍为"—"的任务更新实际文件大小
        self._refresh_completed_sizes()

    def _refresh_completed_sizes(self):
        """周期性检查已完成任务的文件大小，更新表格显示（后备机制）"""
        for row in range(self._table.rowCount()):
            item0 = self._table.item(row, 0)
            if not item0:
                continue
            task_id = item0.data(Qt.UserRole)
            size_item = self._table.item(row, 4)
            if not size_item:
                continue
            # 只处理大小仍为初始值"—"的行
            if size_item.text() == "—":
                # 查找任务
                task = self._dm.get_task(task_id)
                if not task:
                    task = next((t for t in self._pending_tasks if t.task_id == task_id), None)
                if task and task.file_path and os.path.exists(task.file_path):
                    actual_size = os.path.getsize(task.file_path)
                    if actual_size > 0:
                        self._task_file_sizes[task_id] = actual_size
                        size_item.setText(self._format_bytes(actual_size))
                        size_item.setForeground(QColor(51, 51, 51))

    @staticmethod
    def _format_bytes(num_bytes: int) -> str:
        """格式化字节数为可读字符串"""
        if num_bytes <= 0:
            return "0 B"
        for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
            if num_bytes < 1024:
                if unit == 'B':
                    return f"{num_bytes:.0f} {unit}"
                elif unit == 'KB':
                    return f"{num_bytes:.1f} {unit}"
                else:
                    return f"{num_bytes:.2f} {unit}"
            num_bytes /= 1024
        return f"{num_bytes:.2f} PB"

    @staticmethod
    def _format_speed(bytes_per_sec: float) -> str:
        """格式化下载速度"""
        if bytes_per_sec < 1024:
            return f"{bytes_per_sec:.0f} B/s"
        elif bytes_per_sec < 1024 * 1024:
            return f"{bytes_per_sec / 1024:.1f} KB/s"
        else:
            return f"{bytes_per_sec / (1024 * 1024):.1f} MB/s"

    # ------------------------------------------------------------------ #
    #  时间选择
    # ------------------------------------------------------------------ #

    def _set_time_range(self, preset: str):
        now = QDateTime.currentDateTime()
        if preset == "today":
            s = QDateTime(now.date(), QTime(0, 0, 0))
            e = now
        elif preset == "yesterday":
            yd = now.date().addDays(-1)
            s  = QDateTime(yd, QTime(0, 0, 0))
            e  = QDateTime(yd, QTime(23, 59, 59))
        elif preset == "last_1h":
            s = now.addSecs(-3600)
            e = now
        elif preset == "last_24h":
            s = now.addDays(-1)
            e = now
        else:
            return
        self._dt_start.setDateTime(s)
        self._dt_end.setDateTime(e)
    
    def _refresh_preset_combo(self):
        """刷新预设下拉列表"""
        self._preset_combo.clear()
        self._preset_combo.addItem("选择自定义预设...", "")
        for name in sorted(self._time_presets.keys()):
            self._preset_combo.addItem(name, name)
    
    def _on_preset_selected(self, text: str):
        """选择预设时应用时间"""
        if not text or text == "选择自定义预设...":
            return
        
        if text in self._time_presets:
            data = self._time_presets[text]
            start = QDateTime.fromString(data['start'], "yyyy-MM-dd HH:mm:ss")
            end = QDateTime.fromString(data['end'], "yyyy-MM-dd HH:mm:ss")
            
            if start.isValid() and end.isValid():
                self._dt_start.setDateTime(start)
                self._dt_end.setDateTime(end)
                self._log_msg(f"[时间预设] 已应用 '{text}': {data['start']} ~ {data['end']}")
    
    def _manage_time_presets(self):
        """打开时间预设管理对话框"""
        dlg = TimePresetDialog(self, self._time_presets.copy())
        if dlg.exec_() == QDialog.Accepted:
            self._time_presets = dlg.get_presets()
            self._refresh_preset_combo()
            self._save_config()
            self._log_msg(f"[时间预设] 已保存 {len(self._time_presets)} 个预设")

    def _show_help(self):
        """显示使用说明对话框"""
        help_text = """
<h2>四川新数录像批量下载器 - 使用说明</h2>

<h3>一、设备连接</h3>
<ol>
<li>点击「➕ 添加设备」按钮，输入 NVR/DVR 设备信息：
    <ul>
        <li><b>设备名称</b>：自定义名称，用于显示</li>
        <li><b>IP地址</b>：设备的网络地址</li>
        <li><b>HTTP端口</b>：默认 80 或 8000</li>
        <li><b>RTSP端口</b>：默认 554</li>
        <li><b>用户名/密码</b>：设备登录凭证</li>
    </ul>
</li>
<li>支持添加多台设备，同时批量下载</li>
</ol>

<h3>二、通道选择</h3>
<ol>
<li>连接设备后，左侧显示通道树（按设备分组）</li>
<li>勾选需要下载录像的通道</li>
<li>设置下载的<b>时间范围</b>（开始时间~结束时间）</li>
<li>可使用「时间预设」快速选择常用时间段</li>
</ol>

<h3>三、下载任务</h3>
<ol>
<li>点击「📋 加入列表」将选中通道添加到任务列表</li>
<li>点击「▶ 开始下载」执行所有任务</li>
<li>支持右键单个任务：开始/停止/重试/删除</li>
</ol>

<h3>四、下载模式说明</h3>
<ul>
<li><b>ISAPI模式</b>（推荐）：通过 HTTP API 下载，速度快、稳定性好</li>
<li><b>SDK模式</b>：通过海康 SDK 下载，兼容性更广，支持大文件分段下载</li>
</ul>

<h3>五、常见问题</h3>
<ul>
<li><b>连接失败</b>：检查 IP、端口、用户名密码是否正确</li>
<li><b>探测失败</b>：设备可能不支持 ISAPI 下载，请切换到 SDK 模式</li>
<li><b>下载超时</b>：录像时间过长，可调整设置中的超时参数</li>
</ul>

<h3>六、技术支持</h3>
<ul>
<li>官方网站：<a href="http://www.scxs.vip">www.scxs.vip</a></li>
</ul>
"""
        msg = QMessageBox(self)
        msg.setWindowTitle("使用说明")
        msg.setTextFormat(Qt.RichText)
        msg.setText(help_text)
        msg.setStyleSheet("QLabel{min-width: 600px;}")
        msg.exec_()

    def _show_about(self):
        """显示关于对话框"""
        about_text = """
<h2>四川新数录像批量下载器</h2>
<p><b>版本：</b>1.0.0</p>
<p><b>功能：</b>海康威视 NVR/DVR 录像批量下载工具</p>
<hr>
<p>支持功能：</p>
<ul>
    <li>多设备同时连接与下载</li>
    <li>通道批量选择与时间范围设置</li>
    <li>ISAPI / SDK 双下载模式</li>
    <li>大文件自动分段与合并</li>
    <li>OSD 通道名称批量设置</li>
    <li>下载进度实时监控</li>
</ul>
<hr>
<p><b>版权所有：四川新数信息技术有限公司</b></p>
<p>官方网站：<a href="http://www.scxs.vip">www.scxs.vip</a></p>
"""
        msg = QMessageBox(self)
        msg.setWindowTitle("关于软件")
        msg.setTextFormat(Qt.RichText)
        msg.setText(about_text)
        msg.setStyleSheet("QLabel{min-width: 400px;}")
        msg.exec_()

    # ------------------------------------------------------------------ #
    #  下载操作
    # ------------------------------------------------------------------ #

    def _show_download_settings(self):
        """显示下载设置对话框"""
        dlg = DownloadSettingsDialog(
            self,
            total_thread_count=self._total_thread_count,
            per_device_thread_count=self._per_device_thread_count,
            merge_mode=self._merge_mode,
            enable_debug_log=self._enable_debug_log,
            skip_transcode=self._skip_transcode,
            download_mode=self._download_mode,
        )
        if dlg.exec_() == QDialog.Accepted:
            settings = dlg.get_settings()
            self._total_thread_count = settings['total_thread_count']
            self._per_device_thread_count = settings['per_device_thread_count']
            self._merge_mode = settings['merge_mode']
            self._enable_debug_log = settings['enable_debug_log']
            self._skip_transcode = settings['skip_transcode']

            # 更新下载模式
            new_mode = settings.get('download_mode', self._download_mode)
            if new_mode != self._download_mode:
                old_name = "ISAPI" if self._download_mode == "isapi" else "SDK"
                self._download_mode = new_mode
                new_name = "ISAPI" if new_mode == "isapi" else "SDK"
                
                # 同步工具栏下拉框
                idx = 0 if new_mode == "isapi" else 1
                self._mode_combo.blockSignals(True)
                self._mode_combo.setCurrentIndex(idx)
                self._mode_combo.blockSignals(False)
                self._log_msg(f"🔄 下载模式切换: {old_name} → {new_name}")

            # 如果下载管理器未运行，更新线程配置
            if not self._dm._running:
                self._dm.max_concurrent = self._total_thread_count
                self._dm.max_concurrent_per_device = self._per_device_thread_count

            merge_mode_text = {'ultra': '极速', 'fast': '快速', 'standard': '标准'}.get(self._merge_mode, '快速')
            debug_text = "开启" if self._enable_debug_log else "关闭"
            transcode_text = "跳过" if self._skip_transcode else "开启"
            mode_name = "ISAPI" if self._download_mode == "isapi" else "SDK"
            self._log_msg(f"设置已更新: 模式:{mode_name}, 总线程{self._total_thread_count}/每NVR{self._per_device_thread_count}, "
                         f"合并模式:{merge_mode_text}, 调试日志:{debug_text}, 转码:{transcode_text}")

    def _on_download_mode_changed(self, index):
        """工具栏下载模式切换回调"""
        mode = self._mode_combo.currentData()
        if mode == self._download_mode:
            return
        old_mode = self._download_mode
        self._download_mode = mode
        mode_name = "ISAPI" if mode == "isapi" else "SDK"
        old_name = "ISAPI" if old_mode == "isapi" else "SDK"
        self._log_msg(f"🔄 下载模式切换: {old_name} → {mode_name}")
        self._save_config()


    def _pick_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择保存目录", self.download_dir)
        if d:
            self.download_dir = d
            self._dir_label.setText(f"  {d}")
            self._save_config()
            self._log_msg(f"保存目录: {d}")

    def _start_download(self):
        """开始下载：根据当前模式选择不同的下载引擎"""
        if not self._pending_tasks:
            # 兼容旧流程：如果没有待下载任务，尝试从当前选中通道创建
            selected = self._get_selected_channels()
            if not selected:
                QMessageBox.warning(self, "提示", "没有待下载的任务，请先勾选通道并点击「加入列表」")
                return
            # 直接走旧流程
            self._start_download_direct()
            return

        os.makedirs(self.download_dir, exist_ok=True)

        all_tasks = list(self._pending_tasks)
        self._pending_tasks.clear()

        # 记录下载开始时间
        now_ts = time.time()
        for task in all_tasks:
            self._download_start_times[task.task_id] = now_ts

        # 根据模式选择不同的下载引擎
        if self._download_mode == "isapi":
            self._start_isapi_download(all_tasks)
        # elif self._download_mode == "hikload":  # HikLoad模式已禁用
        #     self._start_hikload_download(all_tasks)
        else:
            # SDK模式（默认）
            self._start_sdk_download(all_tasks)

        self._btn_start.setEnabled(False)
        self._btn_stop.setEnabled(True)

        # 启动下载速度刷新定时器
        if not hasattr(self, '_speed_timer') or not self._speed_timer.isActive():
            self._speed_timer = QTimer(self)
            self._speed_timer.timeout.connect(self._update_download_speed)
            self._speed_timer.start(2000)  # 每2秒刷新

        # 统计
        device_keys = set(t.device_id for t in all_tasks)
        channel_count = len(all_tasks)
        mode_name = "ISAPI" if self._download_mode == "isapi" else "SDK"
        self._log_msg(f"▶ 开始下载 [{mode_name}模式] {len(device_keys)}台设备, {channel_count}个通道")
        self.statusBar().showMessage(f"正在下载 {channel_count} 个任务 [{mode_name}模式]...")

    def _start_sdk_download(self, tasks: List[DownloadTask]):
        """SDK模式下载：使用DownloadManager（Java SDK + JNA）"""
        self._dm.add_tasks_batch(tasks)
        self._dm.max_concurrent = self._total_thread_count
        self._dm.max_concurrent_per_device = self._per_device_thread_count
        self._dm.start()
        # 初始化停止事件集合（SDK模式由DM管理）
        if not hasattr(self, '_isapi_stop_events'):
            self._isapi_stop_events = {}

    def _start_isapi_download(self, tasks: List[DownloadTask]):
        """ISAPI模式下载：使用HikvisionISAPI直接HTTP下载（共用同一个表格）
        
        线程控制策略（与SDK模式一致）：
        - 全局信号量：限制总并发下载数（self._total_thread_count）
        - 每设备信号量：限制每台NVR并发数（self._per_device_thread_count）
        """
        if not hasattr(self, '_isapi_stop_events'):
            self._isapi_stop_events = {}

        # 创建全局并发信号量（限制总下载线程数）
        if not hasattr(self, '_isapi_global_sem'):
            self._isapi_global_sem = threading.Semaphore(self._total_thread_count)
        else:
            # 更新信号量容量（如果配置改变则重建）
            current_val = getattr(self._isapi_global_sem, '_initial_value', None)
            if current_val != self._total_thread_count:
                self._isapi_global_sem = threading.Semaphore(self._total_thread_count)
        self._isapi_global_sem._initial_value = self._total_thread_count

        # 按设备分组，为每台设备创建信号量
        if not hasattr(self, '_isapi_device_sems'):
            self._isapi_device_sems = {}
        for task in tasks:
            device_id = task.device_id or "unknown"
            if device_id not in self._isapi_device_sems:
                sem = threading.Semaphore(self._per_device_thread_count)
                sem._initial_value = self._per_device_thread_count
                self._isapi_device_sems[device_id] = sem
            else:
                # 更新信号量容量
                sem = self._isapi_device_sems[device_id]
                old_val = getattr(sem, '_initial_value', None)
                if old_val != self._per_device_thread_count:
                    new_sem = threading.Semaphore(self._per_device_thread_count)
                    new_sem._initial_value = self._per_device_thread_count
                    self._isapi_device_sems[device_id] = new_sem

        for task in tasks:
            config = task.device_config or {}
            stop_event = threading.Event()
            self._isapi_stop_events[task.task_id] = stop_event

            # 更新任务状态为等待中
            task.status = DownloadStatus.PENDING
            task.progress = 0
            self._status_signal.emit(task.task_id)

            # 获取该任务的设备信号量
            device_id = task.device_id or "unknown"
            device_sem = self._isapi_device_sems[device_id]

            t = threading.Thread(
                target=self._isapi_download_worker,
                args=(task, stop_event, self._isapi_global_sem, device_sem),
                name=f"ISAPI-{task.channel_name}",
                daemon=True,
            )
            t.start()

    def _isapi_download_worker(self, task: DownloadTask, stop_event: threading.Event,
                                global_sem: threading.Semaphore, device_sem: threading.Semaphore):
        """ISAPI下载工作线程：下载单个任务并更新表格
        
        Args:
            task: 下载任务
            stop_event: 停止事件
            global_sem: 全局并发信号量（限制总线程数）
            device_sem: 设备并发信号量（限制每NVR线程数）
        """
        from core.nvr_api import create_isapi
        import time as _time

        config = task.device_config or {}
        channel_no = int(task.channel_id) if task.channel_id.isdigit() else 1

        # 先获取全局信号量（等待全局槽位）
        if stop_event.is_set():
            task.status = DownloadStatus.CANCELLED
            self._status_signal.emit(task.task_id)
            return

        global_acquired = global_sem.acquire(timeout=600)
        if not global_acquired:
            task.status = DownloadStatus.FAILED
            task.error_message = "等待全局下载槽位超时"
            self._status_signal.emit(task.task_id)
            self._log_signal.emit(f"✗ {task.channel_name} - 等待全局下载槽位超时")
            return

        try:
            # 再获取设备信号量（等待该NVR的槽位）
            if stop_event.is_set():
                task.status = DownloadStatus.CANCELLED
                self._status_signal.emit(task.task_id)
                return

            device_acquired = device_sem.acquire(timeout=600)
            if not device_acquired:
                task.status = DownloadStatus.FAILED
                task.error_message = "等待设备下载槽位超时"
                self._status_signal.emit(task.task_id)
                self._log_signal.emit(f"✗ {task.channel_name} - 等待设备下载槽位超时")
                return

            try:
                # 获得槽位后才标记为下载中
                task.status = DownloadStatus.DOWNLOADING
                task.progress = 0
                self._status_signal.emit(task.task_id)

                api = create_isapi(config)


                def _progress(pct):
                    task.progress = pct
                    self._progress_signal.emit(task.task_id, pct)

                def _log(msg):
                    self._log_signal.emit(f"[{task.channel_name}] {msg}")

                def _size(size_bytes):
                    """连接成功后立即回调，更新表格中的录像大小"""
                    self._task_file_sizes[task.task_id] = size_bytes
                    self._size_signal.emit(task.task_id, size_bytes)

                t0 = _time.time()

                success, msg = api.download_record_by_time(
                    channel=channel_no,
                    start_time=task.start_time,
                    end_time=task.end_time,
                    save_path=task.file_path,
                    stream_type=1,
                    rtsp_port=config.get('rtsp_port', 554),
                    progress_callback=_progress,
                    log_callback=_log,
                    stop_event=stop_event,
                    size_callback=_size,
                )

                elapsed = _time.time() - t0

                # 更新任务状态
                if success:
                    task.status = DownloadStatus.COMPLETED
                    task.progress = 100
                    # 录像大小在 _cleanup_speed_tracking 中通过主线程统一更新（线程安全）
                    self._log_signal.emit(f"✓ ISAPI下载完成: {task.channel_name} - {msg}, 耗时:{elapsed:.1f}s")
                else:
                    task.status = DownloadStatus.FAILED
                    task.error_message = msg
                    self._log_signal.emit(f"✗ ISAPI下载失败: {task.channel_name} - {msg}")

                # 通知主线程更新表格状态
                self._status_signal.emit(task.task_id)
                # 通知完成
                self._dm.tasks[task.task_id] = task  # 注册到DM以便表格查询
                QTimer.singleShot(0, lambda: self._on_task_done_bg(task.task_id, success, task.file_path, msg))

            except Exception as e:
                task.status = DownloadStatus.FAILED
                task.error_message = str(e)
                self._status_signal.emit(task.task_id)
                self._log_signal.emit(f"✗ ISAPI下载异常: {task.channel_name} - {str(e)}")
                self._dm.tasks[task.task_id] = task
                QTimer.singleShot(0, lambda: self._on_task_done_bg(task.task_id, False, "", str(e)))
            finally:
                # 释放设备信号量
                device_sem.release()
        finally:
            # 释放全局信号量
            global_sem.release()

    def _start_hikload_download(self, tasks: List[DownloadTask]):
        """HikLoad模式下载：使用RTSP+FFmpeg流式下载"""
        from core.hikload_downloader import HikLoadDownloader
        from datetime import datetime
        import threading

        self._log_msg("🚀 开始 HikLoad 批量下载（RTSP+FFmpeg）...")

        # 按设备分组任务
        device_tasks = {}
        for task in tasks:
            device_key = task.device_id or "unknown"
            if device_key not in device_tasks:
                device_tasks[device_key] = []
            device_tasks[device_key].append(task)

        # 为每个设备启动一个HikLoad下载线程
        for device_key, task_list in device_tasks.items():
            if not task_list:
                continue

            # 获取设备配置
            device_config = task_list[0].device_config or {}
            host = device_config.get('host', 'unknown')
            username = device_config.get('username', 'admin')
            password = device_config.get('password', '')

            # 创建停止事件
            stop_event = threading.Event()
            for task in task_list:
                self._isapi_stop_events[task.task_id] = stop_event

            # 启动HikLoad下载线程
            t = threading.Thread(
                target=self._hikload_download_worker,
                args=(host, username, password, task_list, stop_event),
                name=f"HikLoad-{device_key}",
                daemon=True
            )
            t.start()

        self._log_msg(f"✓ 已启动 {len(device_tasks)} 个HikLoad下载线程")

    def _hikload_download_worker(self, host: str, username: str, password: str,
                                   tasks: List[DownloadTask], stop_event: threading.Event):
        """HikLoad下载工作线程：批量下载单个设备的多个通道"""
        from core.hikload_downloader import HikLoadDownloader
        from datetime import datetime

        # 创建HikLoad下载器
        downloader = HikLoadDownloader(
            nvr_ip=host,
            username=username,
            password=password,
            download_path=self.download_dir,
            video_format="mkv",
            use_ffmpeg=True,
            folder_structure="onepercamera"
        )

        # 设置日志回调
        def log_callback(msg):
            self._log_signal.emit(f"[HikLoad-{host}] {msg}")

        # 设置进度回调
        def progress_callback(percent, status):
            # HikLoad批量下载，整体进度
            overall_progress = min(100, max(0, percent))
            for task in tasks:
                task.progress = overall_progress
                self._progress_signal.emit(task.task_id, overall_progress)

        downloader.set_log_callback(log_callback)
        downloader.set_progress_callback(progress_callback)

        # 准备下载参数
        camera_ids = [task.channel_id for task in tasks]
        start_time = tasks[0].start_time
        end_time = tasks[0].end_time

        # 更新所有任务状态为下载中
        for task in tasks:
            task.status = DownloadStatus.DOWNLOADING
            self._status_signal.emit(task.task_id)
            self._download_start_times[task.task_id] = time.time()

        try:
            # 批量下载
            results = downloader.download_videos(
                camera_ids=camera_ids,
                start_time=start_time,
                end_time=end_time,
                concat_videos=True
            )

            # 处理结果
            success_videos = {v['camera_id']: v for v in results['videos']}

            for task in tasks:
                if stop_event.is_set():
                    task.status = DownloadStatus.CANCELLED
                    self._status_signal.emit(task.task_id)
                    continue

                channel_id = task.channel_id
                if channel_id in success_videos:
                    # 下载成功
                    video_info = success_videos[channel_id]
                    task.status = DownloadStatus.COMPLETED
                    task.filepath = video_info['filepath']
                    task.filesize = video_info['size']

                    # 更新文件大小
                    self._task_file_sizes[task.task_id] = video_info['size']
                    self._size_signal.emit(task.task_id, video_info['size'])

                    self._log_signal.emit(f"✓ 通道 {channel_id} 下载完成: {video_info['filepath']}")
                else:
                    # 下载失败
                    task.status = DownloadStatus.FAILED
                    self._log_signal.emit(f"✗ 通道 {channel_id} 下载失败")

                self._status_signal.emit(task.task_id)
                task.progress = 100
                self._progress_signal.emit(task.task_id, 100)

            self._log_signal.emit(f"✓ HikLoad下载完成: 成功 {results['success']} 个，失败 {results['failed']} 个")

        except Exception as e:
            logger.error(f"HikLoad下载异常", exc_info=True)
            self._log_signal.emit(f"✗ HikLoad下载失败: {str(e)}")

            # 所有任务标记为失败
            for task in tasks:
                task.status = DownloadStatus.FAILED
                self._status_signal.emit(task.task_id)
                task.progress = 0
                self._progress_signal.emit(task.task_id, 0)

    def _start_download_direct(self):
        """直接下载（兼容旧流程：不经过加入列表，直接从选中通道下载）"""
        selected = self._get_selected_channels()
        if not selected:
            QMessageBox.warning(self, "提示", "请先勾选至少一个通道")
            return

        start_dt = self._dt_start.dateTime().toPyDateTime()
        end_dt   = self._dt_end.dateTime().toPyDateTime()
        if start_dt >= end_dt:
            QMessageBox.warning(self, "提示", "开始时间必须早于结束时间")
            return

        os.makedirs(self.download_dir, exist_ok=True)

        import uuid
        all_tasks = []
        device_groups = {}

        for ch in selected:
            device = ch.get('device') or self._current_config
            if not device:
                continue
            device_key = f"{device['host']}:{device.get('port', 8000)}"

            if device_key not in self._device_channels:
                QMessageBox.warning(self, "提示",
                    f"设备 {device.get('name', device['host'])} 未连接，请先连接")
                return

            if device_key not in device_groups:
                device_groups[device_key] = {
                    'config': device,
                    'channels': []
                }
            device_groups[device_key]['channels'].append(ch)

        for device_key, group in device_groups.items():
            device_config = group['config']
            for ch in group['channels']:
                task = DownloadTask(
                    task_id      = str(uuid.uuid4()),
                    device_id    = device_key,
                    device_name  = device_config.get('name', device_config['host']),
                    device_config = device_config,
                    channel_id   = str(ch.get('no', ch.get('id', '1'))),
                    channel_name = ch.get('name', f"通道{ch.get('id','?')}"),
                    start_time   = start_dt,
                    end_time     = end_dt,
                    save_dir     = self.download_dir,
                    merge_mode   = getattr(self, '_merge_mode', 'standard'),
                    enable_debug_log = getattr(self, '_enable_debug_log', False),
                    skip_transcode = getattr(self, '_skip_transcode', True),
                )
                all_tasks.append(task)
                self._add_row(task)

        # 记录下载开始时间
        now_ts = time.time()
        for task in all_tasks:
            self._download_start_times[task.task_id] = now_ts

        # 根据模式选择下载引擎
        if self._download_mode == "isapi":
            self._start_isapi_download(all_tasks)
        else:
            self._dm.add_tasks_batch(all_tasks)
            self._dm.max_concurrent = self._total_thread_count
            self._dm.max_concurrent_per_device = self._per_device_thread_count
            self._dm.start()

        self._btn_start.setEnabled(False)
        self._btn_stop.setEnabled(True)

        # 启动下载速度刷新定时器
        if not hasattr(self, '_speed_timer') or not self._speed_timer.isActive():
            self._speed_timer = QTimer(self)
            self._speed_timer.timeout.connect(self._update_download_speed)
            self._speed_timer.start(2000)

        device_count = len(device_groups)
        channel_count = len(selected)
        mode_name = "ISAPI" if self._download_mode == "isapi" else "SDK"
        self._log_msg(f"▶ 开始下载 [{mode_name}模式] {device_count}台设备, {channel_count}个通道, 总线程{self._total_thread_count}/每NVR{self._per_device_thread_count}")

    def _stop_download(self):
        # 停止SDK下载管理器
        self._dm.stop()

        # 停止ISAPI下载任务
        if hasattr(self, '_isapi_stop_events'):
            for task_id, stop_event in self._isapi_stop_events.items():
                stop_event.set()
            self._isapi_stop_events.clear()

        self._btn_start.setEnabled(True)
        self._btn_stop.setEnabled(False)
        # 停止速度刷新定时器
        if hasattr(self, '_speed_timer') and self._speed_timer.isActive():
            self._speed_timer.stop()
        self._speed_label.setText("下载速度: --")
        self._speed_label.setStyleSheet("font-size: 13px; color: #999;")
        self._log_msg("■ 已停止下载")
        self.statusBar().showMessage("已停止下载")

    def _on_table_context_menu(self, pos):
        """下载表格右键菜单"""
        row = self._table.rowAt(pos.y())
        if row < 0:
            return

        item0 = self._table.item(row, 0)
        if not item0:
            return
        task_id = item0.data(Qt.UserRole)
        # 先从 DownloadManager 查找，找不到再从 _pending_tasks 查找
        task = self._dm.get_task(task_id)
        if not task:
            task = next((t for t in self._pending_tasks if t.task_id == task_id), None)
        if not task:
            return

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { font-size: 13px; padding: 4px; }
            QMenu::item { padding: 6px 24px; }
            QMenu::item:disabled { color: #aaa; }
        """)

        # 开始下载：仅对等待中(PENDING)的任务可用
        can_start = task.status == DownloadStatus.PENDING
        action_start = menu.addAction("▶ 开始下载此任务")
        action_start.setEnabled(can_start)
        action_start.triggered.connect(lambda checked, t=task: self._start_single_task(t))

        # 重新下载：仅对失败或已取消的任务可用
        can_retry = task.status in (DownloadStatus.FAILED, DownloadStatus.CANCELLED)
        action_retry = menu.addAction("🔄 重新下载")
        action_retry.setEnabled(can_retry)
        action_retry.triggered.connect(lambda checked, t=task: self._retry_single_task(t))

        menu.addSeparator()

        # 停止按钮：仅对等待中或下载中的任务可用
        can_stop = task.status in (DownloadStatus.PENDING, DownloadStatus.DOWNLOADING, DownloadStatus.MERGING)
        action_stop = menu.addAction("⏹ 停止此任务")
        action_stop.setEnabled(can_stop)
        action_stop.triggered.connect(lambda checked, tid=task_id, t=task: self._stop_single_task(tid, t))

        menu.addSeparator()

        # 删除按钮：仅对非下载中的任务可用
        can_delete = task.status not in (DownloadStatus.DOWNLOADING, DownloadStatus.MERGING)
        action_delete = menu.addAction("🗑 删除此任务")
        action_delete.setEnabled(can_delete)
        action_delete.triggered.connect(lambda checked, tid=task_id, t=task: self._delete_single_task(tid, t))

        menu.addSeparator()

        # 打开文件位置：仅对已完成的任务可用
        can_open = task.status == DownloadStatus.COMPLETED and task.file_path and os.path.exists(task.file_path)
        action_open = menu.addAction("📁 打开文件位置")
        action_open.setEnabled(can_open)
        action_open.triggered.connect(lambda checked, fp=task.file_path: self._open_file_location(fp))

        menu.exec_(self._table.viewport().mapToGlobal(pos))

    def _stop_single_task(self, task_id: str, task: DownloadTask):
        """停止单个下载任务"""
        # 从待下载列表中移除（如果还在的话）
        self._pending_tasks = [t for t in self._pending_tasks if t.task_id != task_id]

        if task.status == DownloadStatus.PENDING:
            # SDK模式取消
            self._dm.cancel_task(task_id)
            self._status_signal.emit(task_id)
            self._log_msg(f"⏹ 已停止任务: {task.channel_name}")
        elif task.status == DownloadStatus.DOWNLOADING:
            # 下载中：通过ISAPI停止事件标记停止
            if task_id in self._isapi_stop_events:
                self._isapi_stop_events[task_id].set()
            # SDK模式也标记取消
            self._dm.cancel_task_downloading(task_id)
            self._status_signal.emit(task_id)
            self._log_msg(f"⏹ 正在停止任务: {task.channel_name}")

        self._update_stats()

    def _delete_single_task(self, task_id: str, task: DownloadTask):
        """删除单个任务（从列表中移除）"""
        # 如果正在下载，先停止
        if task.status == DownloadStatus.DOWNLOADING:
            if task_id in self._isapi_stop_events:
                self._isapi_stop_events[task_id].set()
            self._dm.cancel_task_downloading(task_id)

        # 从待下载列表中移除
        self._pending_tasks = [t for t in self._pending_tasks if t.task_id != task_id]

        # 从DownloadManager中移除
        self._dm.remove_task(task_id)

        # 从表格中删除行
        row = self._find_row(task_id)
        if row >= 0:
            self._table.removeRow(row)


        # 清理相关跟踪数据
        self._task_file_sizes.pop(task_id, None)
        self._download_start_times.pop(task_id, None)
        self._isapi_stop_events.pop(task_id, None)

        self._log_msg(f"🗑 已删除任务: {task.channel_name}")
        self._update_stats()

        # 如果表格为空，启用开始按钮但禁用停止按钮
        if self._table.rowCount() == 0:
            self._btn_start.setEnabled(len(self._pending_tasks) > 0)
            self._btn_stop.setEnabled(False)

    def _start_single_task(self, task: DownloadTask):
        """开始下载单个任务"""
        if task.status != DownloadStatus.PENDING:
            return
        
        # 根据下载模式启动单个任务
        if self._download_mode == "isapi":
            self._start_isapi_download([task])
        else:
            self._dm.add_task(task)
            self._dm.start()
        
        self._btn_start.setEnabled(False)
        self._btn_stop.setEnabled(True)
        self._log_msg(f"▶ 开始下载: {task.channel_name}")

    def _retry_single_task(self, task: DownloadTask):
        """重新下载失败/取消的任务"""
        if task.status not in (DownloadStatus.FAILED, DownloadStatus.CANCELLED):
            return
        
        # 重置任务状态
        task.status = DownloadStatus.PENDING
        task.progress = 0
        task.error_message = ""
        
        # 更新表格显示
        self._status_signal.emit(task.task_id)
        
        # 启动下载
        if self._download_mode == "isapi":
            self._start_isapi_download([task])
        else:
            # SDK模式：重新添加到队列
            self._dm.add_task(task)
            self._dm.start()
        
        self._btn_start.setEnabled(False)
        self._btn_stop.setEnabled(True)
        self._log_msg(f"🔄 重新下载: {task.channel_name}")

    def _open_file_location(self, file_path: str):
        """打开文件所在目录"""
        if not file_path or not os.path.exists(file_path):
            QMessageBox.warning(self, "提示", "文件不存在")
            return
        
        # Windows: 选中文件并打开目录
        import subprocess
        try:
            subprocess.run(['explorer', '/select,', os.path.abspath(file_path)], check=False)
        except Exception as e:
            # 回退：只打开目录
            try:
                os.startfile(os.path.dirname(file_path))
            except Exception:
                QMessageBox.warning(self, "提示", f"无法打开目录: {e}")

    def _clear_completed(self):
        self._dm.clear_completed()
        self._table.setRowCount(0)
        for t in self._dm.get_all_tasks():
            self._add_row(t)
        self._update_stats()

    # ------------------------------------------------------------------ #
    #  任务表格
    # ------------------------------------------------------------------ #

    def _add_row(self, task: DownloadTask, initial_size: int = 0):
        """
        添加一行任务到表格。

        Args:
            task:         下载任务
            initial_size: 初始录像大小（字节）；>0=直接显示大小，0=显示"探测中..."
        """
        row = self._table.rowCount()
        self._table.insertRow(row)

        def item(text):
            i = QTableWidgetItem(text)
            i.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            return i

        self._table.setItem(row, 0, item(task.device_name))
        self._table.setItem(row, 1, item(task.channel_name))
        self._table.setItem(row, 2, item(task.start_time.strftime("%m-%d %H:%M")))
        self._table.setItem(row, 3, item(task.end_time.strftime("%m-%d %H:%M")))

        # 录像大小（第4列）
        # initial_size: >0=实测大小(深灰)  0=探测中(灰)  -1=探测失败(红)  -2=离线(橙)
        if initial_size > 0:
            size_item = item(self._format_bytes(initial_size))
            size_item.setForeground(QColor(51, 51, 51))
        elif initial_size == -1:
            size_item = item("探测失败")
            size_item.setForeground(QColor(244, 67, 54))  # 红色
        elif initial_size == -2:
            size_item = item("离线")
            size_item.setForeground(QColor(255, 152, 0))  # 橙色
        else:
            size_item = item("探测中...")
            size_item.setForeground(QColor(160, 160, 160))
        self._table.setItem(row, 4, size_item)

        status_item = QTableWidgetItem(STATUS_TEXT[task.status])
        status_item.setForeground(STATUS_COLORS[task.status])
        self._table.setItem(row, 5, status_item)

        # 下载进度条（第6列）
        download_bar = QProgressBar()
        download_bar.setRange(0, 100)
        download_bar.setValue(task.progress)
        download_bar.setTextVisible(True)
        self._table.setCellWidget(row, 6, download_bar)

        # 在第0列存task_id
        self._table.item(row, 0).setData(Qt.UserRole, task.task_id)

    def _find_row(self, task_id: str) -> int:
        for r in range(self._table.rowCount()):
            item = self._table.item(r, 0)
            if item and item.data(Qt.UserRole) == task_id:
                return r
        return -1


    # ------------------------------------------------------------------ #
    #  回调槽（主线程）
    # ------------------------------------------------------------------ #

    def _update_size_in_table(self, task_id: str, size_bytes: int, is_estimate: bool = False):
        """
        更新表格中的录像大小列。

        Args:
            task_id:    任务ID
            size_bytes: 文件大小（字节）；若 <= 0 且 is_estimate=True 则根据码率×时长估算
            is_estimate: True=估算大小（橙色显示"估算: X MB"），False=实测大小（深灰显示）
        """
        row = self._find_row(task_id)
        if row < 0:
            return

        item = self._table.item(row, 4)
        if not item:
            return

        # ISAPI探测成功：显示实测大小
        if size_bytes > 0:
            item.setText(self._format_bytes(size_bytes))
            item.setForeground(QColor(51, 51, 51))
            return

        # 探测失败/未知：保持"探测中..."状态（用户可开始下载，实际大小会在下载时更新）
        if not is_estimate:
            return  # size <= 0 且非估算，保持原样

        # 估算大小：码率 × 时长
        # 查找任务的设备配置和通道号，尝试从已缓存的通道信息获取码率
        task = self._dm.get_task(task_id)
        if not task:
            task = next((t for t in self._pending_tasks if t.task_id == task_id), None)
        if not task:
            return

        device_key = task.device_id
        ch_no = int(task.channel_id) if task.channel_id.isdigit() else 1
        channels = self._device_channels.get(device_key, [])
        ch_info = next((c for c in channels if int(c.get('no', c.get('id', 0))) == ch_no), None)

        bitrate_kbps = 0
        if ch_info and 'bitrate_kbps' in ch_info:
            bitrate_kbps = ch_info['bitrate_kbps']

        if bitrate_kbps <= 0:
            # 无法估算，保持"探测中..."
            return

        duration_sec = (task.end_time - task.start_time).total_seconds()
        if duration_sec <= 0:
            return

        estimated_bytes = int(bitrate_kbps * 1000 / 8 * duration_sec)
        size_str = f"估算: {self._format_bytes(estimated_bytes)}"
        item.setText(size_str)
        item.setForeground(QColor(255, 152, 0))  # 橙色（与"探测中"灰色区分）


    def _on_progress_ui(self, task_id: str, progress: int):
        """下载进度更新"""
        row = self._find_row(task_id)
        if row >= 0:
            bar = self._table.cellWidget(row, 6)  # 第6列：下载进度
            if bar:
                bar.setValue(progress)
                bar.setFormat(f"{progress}%")

    def _on_status_ui(self, task_id: str):
        task = self._dm.get_task(task_id)
        if not task:
            return
        row = self._find_row(task_id)
        if row >= 0:
            si = self._table.item(row, 5)
            if si:
                si.setText(STATUS_TEXT[task.status])
                si.setForeground(STATUS_COLORS[task.status])
        self._update_stats()
        # 更新状态栏
        self._update_status_bar_from_tasks()

    def _on_task_done_bg(self, task_id: str, success: bool, file_path: str, error_message: str):
        """后台线程回调 → 转发到主线程日志"""
        task = self._dm.get_task(task_id)
        if not task:
            return
        icon = "✅" if success else "❌"
        msg  = f"{icon} {task.channel_name}: {STATUS_TEXT[task.status]}"
        if error_message:
            msg += f"  (错误: {error_message})"
        # 失败时同时输出到运行日志和下载日志，确保能看到错误信息
        QTimer.singleShot(0, lambda: self._log_msg(msg))
        QTimer.singleShot(0, lambda: self._log_download(msg))
        QTimer.singleShot(0, self._update_stats)
        # 清理速度跟踪数据
        QTimer.singleShot(0, lambda: self._cleanup_speed_tracking(task_id, success, file_path))

    def _cleanup_speed_tracking(self, task_id: str, success: bool, file_path: str):
        """清理单个任务的速度跟踪数据，并在主线程更新录像大小"""
        self._download_start_times.pop(task_id, None)

        # 如果下载成功且有实际文件，更新实际大小和表格显示
        # 优先使用传入的file_path，如果不存在则从task对象获取
        actual_path = file_path
        if success and actual_path and not os.path.exists(actual_path):
            # 传入的路径不存在，尝试从DM/task对象获取最新路径
            task = self._dm.get_task(task_id)
            if task and task.file_path and os.path.exists(task.file_path):
                actual_path = task.file_path

        if success and actual_path and os.path.exists(actual_path):
            actual_size = os.path.getsize(actual_path)
            if actual_size > 0:
                self._task_file_sizes[task_id] = actual_size
                self._update_size_in_table(task_id, actual_size)

        # 检查是否所有任务都已完成
        if not self._download_start_times:
            if hasattr(self, '_speed_timer') and self._speed_timer.isActive():
                self._speed_timer.stop()
            self._speed_label.setText("下载速度: --")
            self._speed_label.setStyleSheet("font-size: 13px; color: #999;")
        # 刷新磁盘信息
        self._update_disk_info()

    def _update_stats(self):
        tasks     = self._dm.get_all_tasks()
        total     = len(tasks)
        completed = sum(1 for t in tasks if t.status == DownloadStatus.COMPLETED)
        failed    = sum(1 for t in tasks if t.status == DownloadStatus.FAILED)
        # 计算已完成任务的实际文件大小
        completed_size = 0
        for t in tasks:
            if t.status == DownloadStatus.COMPLETED and t.file_path and os.path.exists(t.file_path):
                completed_size += os.path.getsize(t.file_path)
        size_str = f" ({self._format_bytes(completed_size)})" if completed_size > 0 else ""
        # 更新下载任务信息面板中的统计标签
        self._task_stats_label.setText(f"任务: {total} | 完成: {completed}{size_str} | 失败: {failed}")

    def _update_status_bar_from_tasks(self):
        """根据所有任务状态更新底部状态栏"""
        tasks = self._dm.get_all_tasks()
        if not tasks:
            self.statusBar().showMessage("就绪")
            return

        downloading = sum(1 for t in tasks if t.status == DownloadStatus.DOWNLOADING)
        pending = sum(1 for t in tasks if t.status == DownloadStatus.PENDING)
        completed = sum(1 for t in tasks if t.status == DownloadStatus.COMPLETED)
        failed = sum(1 for t in tasks if t.status == DownloadStatus.FAILED)

        if downloading > 0:
            self.statusBar().showMessage(f"正在下载 {downloading} 个任务...")
        elif pending > 0:
            self.statusBar().showMessage(f"等待下载 {pending} 个任务")
        elif completed > 0 and failed == 0:
            self.statusBar().showMessage(f"全部完成 ({completed} 个任务)")
        elif failed > 0:
            self.statusBar().showMessage(f"下载完成: {completed} 成功, {failed} 失败")
        else:
            self.statusBar().showMessage("就绪")

    # ------------------------------------------------------------------ #
    #  日志
    # ------------------------------------------------------------------ #

    def _switch_log_view(self, mode: str):
        """切换日志视图模式（互斥）"""
        self._log_view_mode = mode
        if mode == "run":
            self._radio_run_log.setChecked(True)
            self._radio_download_log.setChecked(False)
            self._log_label.setText("运行日志:")
            # 显示运行日志内容
            self._log_text.clear()
            for line in self._run_log_buffer:
                self._log_text.append(line)
        else:
            self._radio_run_log.setChecked(False)
            self._radio_download_log.setChecked(True)
            self._log_label.setText("下载日志:")
            # 显示下载日志内容
            self._log_text.clear()
            for line in self._download_log_buffer:
                self._log_text.append(line)
    
    def _clear_all_logs(self):
        """清除所有日志"""
        self._run_log_buffer.clear()
        self._download_log_buffer.clear()
        self._log_text.clear()
    
    def _log_msg(self, msg: str):
        """添加运行日志（过滤进度信息）"""
        # 过滤掉进度信息，只显示重要信息
        if "Progress:" in msg and "%" in msg:
            return
        ts = datetime.now().strftime("%H:%M:%S")
        log_line = f"[{ts}] {msg}"
        self._run_log_buffer.append(log_line)
        # 如果当前显示的是运行日志，则实时更新
        if self._log_view_mode == "run":
            self._log_text.append(log_line)
    
    def _log_download(self, msg: str):
        """添加下载日志（只显示关键信息）"""
        # 只保留关键日志关键词
        key_keywords = [
            "开始下载", "下载完成", "下载失败", "下载取消",
            "开始合并", "合并完成", "合并失败",
            "分段", "转码", "清理", "临时",
            "✓", "✗", "→", "▼"
        ]
        
        # 过滤掉纯进度信息和Java详细输出
        skip_keywords = [
            "Progress:", "[Java]", "[DownloadManager DEBUG]",
            "收到日志:", "_gui_log 被调用", "回调类型"
        ]
        
        # 如果包含跳过关键词，不显示
        for skip in skip_keywords:
            if skip in msg:
                return
        
        # 只显示关键信息
        has_key = any(kw in msg for kw in key_keywords)
        if not has_key:
            return
        
        ts = datetime.now().strftime("%H:%M:%S")
        log_line = f"[{ts}] {msg}"
        self._download_log_buffer.append(log_line)
        # 如果当前显示的是下载日志，则实时更新
        if self._log_view_mode == "download":
            self._log_text.append(log_line)

    def _export_log(self):
        """导出日志到文件"""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"hikvision_download_log_{ts}.txt"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出日志", default_name, "文本文件 (*.txt);;所有文件 (*)"
        )
        if path:
            try:
                with open(path, 'w', encoding='utf-8') as f:
                    f.write("=" * 60 + "\n")
                    f.write("运行日志\n")
                    f.write("=" * 60 + "\n\n")
                    f.write("\n".join(self._run_log_buffer))
                    f.write("\n\n" + "=" * 60 + "\n")
                    f.write("下载日志\n")
                    f.write("=" * 60 + "\n\n")
                    f.write("\n".join(self._download_log_buffer))
                self._log_msg(f"✅ 日志已导出到: {path}")
                QMessageBox.information(self, "导出成功", f"日志已导出到:\n{path}")
            except Exception as e:
                self._log_msg(f"❌ 导出日志失败: {e}")
                QMessageBox.critical(self, "导出失败", f"导出日志失败:\n{e}")

    # ------------------------------------------------------------------ #
    #  配置存取
    # ------------------------------------------------------------------ #

    def _load_config(self):
        path = os.path.join(os.path.expanduser("~"), ".hikvision_downloader", "config.json")
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                    self.devices      = cfg.get('devices', [])
                    self.download_dir = cfg.get('download_dir', self.download_dir)
                    self._time_presets = cfg.get('time_presets', {})
                    self._download_mode = cfg.get('download_mode', 'isapi')
            except Exception as e:
                print(f"加载配置失败: {e}")

    def _save_config(self):
        d = os.path.join(os.path.expanduser("~"), ".hikvision_downloader")
        os.makedirs(d, exist_ok=True)
        path = os.path.join(d, "config.json")
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump({
                    'devices':      self.devices,
                    'download_dir': self.download_dir,
                    'time_presets': self._time_presets,
                    'download_mode': self._download_mode,
                }, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存配置失败: {e}")

    def _import_config(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入配置", "", "JSON (*.json)")
        if not path:
            return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            self.devices      = cfg.get('devices', self.devices)
            self.download_dir = cfg.get('download_dir', self.download_dir)
            self._save_config()
            self._refresh_device_list()
            self._log_msg("配置导入成功")
        except Exception as e:
            QMessageBox.warning(self, "导入失败", str(e))

    def _export_config(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "导出配置", "hikvision_config.json", "JSON (*.json)"
        )
        if not path:
            return
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump({'devices': self.devices, 'download_dir': self.download_dir},
                          f, ensure_ascii=False, indent=2)
            self._log_msg("配置导出成功")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", str(e))

    def _import_device_list(self):
        """从CSV文件导入设备列表"""
        path, _ = QFileDialog.getOpenFileName(
            self, "导入设备列表", "", 
            "CSV文件 (*.csv);;文本文件 (*.txt);;所有文件 (*)"
        )
        if not path:
            return
        
        try:
            imported_devices = []
            errors = []
            
            with open(path, 'r', encoding='utf-8-sig') as f:  # utf-8-sig 处理BOM
                # 尝试检测是否是CSV格式
                first_line = f.readline().strip()
                f.seek(0)
                
                # 检查是否是CSV头
                is_csv = ',' in first_line or '，' in first_line
                
                if is_csv and ('名称' in first_line or 'name' in first_line.lower() or 'IP' in first_line):
                    # 有表头的CSV
                    import csv
                    reader = csv.DictReader(f)
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            device = self._parse_device_row(row, row_num)
                            if device:
                                imported_devices.append(device)
                        except Exception as e:
                            errors.append(f"第{row_num}行: {e}")
                else:
                    # 无表头的简单格式：每行一个设备，格式为 name,host,port,username,password
                    for row_num, line in enumerate(f, start=1):
                        line = line.strip()
                        if not line or line.startswith('#'):
                            continue
                        
                        parts = line.replace('，', ',').split(',')
                        if len(parts) >= 2:
                            device = {
                                'name': parts[0].strip() or parts[1].strip(),
                                'host': parts[1].strip(),
                                'port': int(parts[2].strip()) if len(parts) > 2 and parts[2].strip() else 8000,
                                'http_port': int(parts[3].strip()) if len(parts) > 3 and parts[3].strip() else 80,
                                'username': parts[4].strip() if len(parts) > 4 else 'admin',
                                'password': parts[5].strip() if len(parts) > 5 else ''
                            }
                            imported_devices.append(device)
                        else:
                            errors.append(f"第{row_num}行: 格式错误")
            
            if imported_devices:
                # 询问是追加还是替换
                reply = QMessageBox.question(
                    self, "导入设备",
                    f"成功解析 {len(imported_devices)} 个设备\n{len(errors)} 个错误\n\n是追加到现有列表还是替换？",
                    QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                    QMessageBox.Yes
                )
                
                if reply == QMessageBox.Cancel:
                    return
                elif reply == QMessageBox.No:
                    # 替换
                    self.devices = imported_devices
                else:
                    # 追加
                    self.devices.extend(imported_devices)
                
                self._save_config()
                self._refresh_device_list()
                self._log_msg(f"✅ 成功导入 {len(imported_devices)} 个设备")
            
            if errors:
                error_msg = "\n".join(errors[:10])
                if len(errors) > 10:
                    error_msg += f"\n...还有 {len(errors) - 10} 个错误"
                QMessageBox.warning(self, "导入警告", f"部分行导入失败:\n{error_msg}")
                
        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"导入设备列表失败:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def _parse_device_row(self, row: dict, row_num: int) -> dict:
        """解析CSV行数据为设备配置"""
        # 支持多种列名变体
        name = row.get('名称') or row.get('name') or row.get('设备名称') or row.get('Name')
        host = row.get('IP地址') or row.get('host') or row.get('IP') or row.get('ip')
        
        if not host:
            raise ValueError("缺少IP地址")
        
        # 端口
        port_str = row.get('SDK端口') or row.get('port') or row.get('端口') or '8000'
        try:
            port = int(port_str)
        except:
            port = 8000
        
        # HTTP端口
        http_port_str = row.get('HTTP端口') or row.get('http_port') or '80'
        try:
            http_port = int(http_port_str)
        except:
            http_port = 80
        
        # 用户名密码
        username = row.get('用户名') or row.get('username') or row.get('Username') or 'admin'
        password = row.get('密码') or row.get('password') or row.get('Password') or ''
        
        return {
            'name': name or host,
            'host': host,
            'port': port,
            'http_port': http_port,
            'username': username,
            'password': password
        }

    def _export_device_list(self):
        """导出设备列表到CSV文件"""
        if not self.devices:
            QMessageBox.information(self, "提示", "当前没有设备可导出")
            return
        
        path, _ = QFileDialog.getSaveFileName(
            self, "导出设备列表", "hikvision_devices.csv", 
            "CSV文件 (*.csv);;文本文件 (*.txt)"
        )
        if not path:
            return
        
        try:
            import csv
            
            with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                # 写入UTF-8 BOM，让Excel正确识别中文
                fieldnames = ['名称', 'IP地址', 'SDK端口', 'HTTP端口', '用户名', '密码']
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                
                writer.writeheader()
                for device in self.devices:
                    writer.writerow({
                        '名称': device.get('name', ''),
                        'IP地址': device.get('host', ''),
                        'SDK端口': device.get('port', 8000),
                        'HTTP端口': device.get('http_port', 80),
                        '用户名': device.get('username', 'admin'),
                        '密码': device.get('password', '')
                    })
            
            self._log_msg(f"✅ 设备列表已导出到: {path}")
            QMessageBox.information(self, "导出成功", f"成功导出 {len(self.devices)} 个设备到:\n{path}")
            
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出设备列表失败:\n{str(e)}")

    # ------------------------------------------------------------------ #
    #  初始化设备列表
    # ------------------------------------------------------------------ #

    def showEvent(self, event):
        super().showEvent(event)
        self._refresh_device_list()
        if self.devices:
            self._device_list.setCurrentRow(0)

    def closeEvent(self, event):
        self._dm.stop()
        # 停止ISAPI下载任务
        if hasattr(self, '_isapi_stop_events'):
            for task_id, stop_event in self._isapi_stop_events.items():
                stop_event.set()
            self._isapi_stop_events.clear()
        self._save_config()
        event.accept()

